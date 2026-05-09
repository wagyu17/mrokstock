"""
update_plan.py  ─ トレーニング計画 進捗管理スクリプト
────────────────────────────────────────────────────────
【目標】
  2026年9月 部内10000m選考会: 32:30以内（理想32:00）
  2026年10月 箱根駅伝予選会: 出走権獲得

【使い方】
  更新.bat に組み込み済み。単独実行も可能:
  > python update_plan.py

【自動更新内容】
  - 「フェーズ計画」シート: 計画概要（固定）
  - 「週次管理」シート: 週ごとの目標 vs 実績（自動集計）
    ・実走行距離は Garminデータ から自動取得
    ・高強度セッション数は心拍Zone4+ から自動カウント
    ・メモ欄の手入力は更新後も保持される

【設定変更】
  選考会日程が確定したら RACE_DATE を更新してください。
────────────────────────────────────────────────────────
"""
import sys, os
from datetime import datetime, timedelta
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════
#  ★ シーズン設定（変更が必要な場合はここだけ修正）
# ══════════════════════════════════════════════════════
SEASON_START  = datetime(2026, 4, 7)   # Week 1 の月曜日（移行週）
RACE_DATE     = datetime(2026, 9, 19)  # 部内選考会【最早候補: 9/19】★確定次第更新
#               datetime(2026, 9, 26)  # 候補2: 9/26
#               datetime(2026, 10, 3)  # 候補3: 10/3
NEAR_RACE     = datetime(2026, 5, 10)  # 東海大学記録会 10000m
TARGET_TIME   = '32:30（理想32:00）'
MAX_HR        = 187

# ══════════════════════════════════════════════════════
#  フェーズ定義
#  ※ W1 は現行週（移行週）、実質Phase1はW2から
# ══════════════════════════════════════════════════════
PHASES = [
    {'id': 1, 'name': '再構築＆記録会準備', 'weeks': (1,  5),  'color': 'BDE0FB', 'fg': '1D3557'},
    {'id': 2, 'name': '有酸素発展期',        'weeks': (6,  12), 'color': 'AEF0D4', 'fg': '2D6A4F'},
    {'id': 3, 'name': '特異的強化期',        'weeks': (13, 18), 'color': 'FFCBA4', 'fg': '9B3B00'},
    {'id': 4, 'name': 'テーパー期',          'weeks': (19, 24), 'color': 'FFADAD', 'fg': '7D0000'},
]
# 選考会候補日ごとの本番週番号（SEASON_START=4/7基準）
# W24(9/15-9/21): 9/19選考 → 本番週
# W25(9/22-9/28): 9/26選考 → 本番週
# W26(9/29-10/5): 10/3選考 → 本番週
RACE_WEEK = {
    datetime(2026, 9, 19): 24,
    datetime(2026, 9, 26): 25,
    datetime(2026, 10, 3): 26,
}

# ══════════════════════════════════════════════════════
#  週次計画テーブル
#  (目標距離km, 水曜ポイント練習[AM→PM], 土曜ポイント練習[AM→PM], 備考)
#  ダブルスレッショルド: 水曜・土曜それぞれAM+PMの2部構成
# ══════════════════════════════════════════════════════
WEEKLY_PLAN = {
    # ─ W1: 移行週（4/7-4/13）────────────────────────
    1:  (55,
         'イージーラン 10km | PM: 電動TM LT 5km @ 3:45',
         'ペース走 5km @ 3:50/km | PM: エリプティカル',
         '現行週・慎重にスタート'),
    # ─ Phase 1: 再構築＆東海大記録会準備（W2-W5）────
    2:  (65,
         '4×2000m @ 3:40/km（AM） | 16×400m @ 3:20/km（PM・TM）',
         '6×1000m @ 3:30/km（AM） | 10×500m @ 3:20/km（PM・TM）',
         '第1週・ダブルLT適応'),
    3:  (75,
         '5×2000m @ 3:38/km（AM） | 20×400m @ 3:18/km（PM・TM）',
         '8×1000m @ 3:28/km（AM） | 12×500m @ 3:18/km（PM・TM）',
         '最大負荷週'),
    4:  (75,
         '4×2000m @ 3:36/km（AM） | 10×400m @ 3:15/km（PM）',
         '5000m＋3000m @ 3:23/km（AM・記録会ペース確認）',
         '記録会前ピーク'),
    5:  (45,
         '4×1000m＋2×400m（軽め・AM）',
         '【東海大学記録会】10000m 目標33:50（3:23/km）',
         '★5/10 東海大学記録会'),
    # ─ Phase 2: 有酸素発展期（W6-W12）──────────────
    6:  (80,
         '5×2000m @ 3:35/km（AM） | 20×400m @ 3:15/km（PM・TM）',
         '8×1000m @ 3:25/km（AM） | 12×500m @ 3:15/km（PM）',
         '記録会後リセット'),
    7:  (88,
         '3000m×4 @ 3:33/km（AM） | 10×400m @ 3:12/km（PM）',
         '1000m×8 @ 3:20/km（AM） | 800m×8 @ 3:08/km（PM）',
         ''),
    8:  (95,
         '3000m×4 @ 3:30/km（AM） | 10×400m @ 3:10/km（PM）',
         '1200m×6 @ 3:18/km（AM） | 800m×10 @ 3:05/km（PM）',
         ''),
    9:  (100,
         '5000m＋3000m＋2000m @ 3:28/km（AM） | LT5km（PM・TM）',
         '1000m×10 @ 3:18/km（AM） | 600m×8 @ 3:05/km（PM）',
         ''),
    10: (100,
         '3000m×4 @ 3:27/km（AM） | 10×400m @ 3:08/km（PM）',
         '800m×12 @ 3:08/km（AM） | 200m×16 @ 3:00/km（PM）',
         ''),
    11: (105,
         '3000m×5 @ 3:25/km（AM） | 10×400m @ 3:07/km（PM）',
         '1000m×10 @ 3:15/km（AM） | 600m×8 @ 3:03/km（PM）',
         ''),
    12: (85,
         '2000m×3 @ 3:32/km（回復週） | LT4km（PM）',
         '5km ペース走 @ 3:35/km',
         '回復週'),
    # ─ Phase 3: 特異的強化期（W13-W18）─────────────
    13: (100,
         '3000m×4 @ 3:22/km（AM） | 10×400m @ 3:05/km（PM）',
         '1000m×8 @ 3:12/km（AM） | 600m×10 @ 3:00/km（PM）',
         ''),
    14: (108,
         '5000m×2 @ 3:22/km（AM） | 10×400m @ 3:03/km（PM）',
         '800m×12 @ 3:05/km（AM） | 400m×10 @ 2:58/km（PM）',
         ''),
    15: (110,
         '5000m＋3000m＋2000m @ 3:20→3:15/km（AM） | 8×400m（PM）',
         '1000m×10 @ 3:10/km（AM） | 600m×8 @ 3:00/km（PM）',
         ''),
    16: (95,
         '2000m×3 @ 3:25/km（回復週） | LT4km（PM）',
         '1000m×5 @ 3:12/km',
         '回復週'),
    17: (108,
         '【中間TT】10000m 目標33:00（AM）',
         '400m×12 @ 2:58/km（r60s）',
         '★中間タイムトライアル'),
    18: (85,
         '5000m @ 3:15/km（AM・レースペース確認）',
         '5000m @ 3:18/km',
         'テーパー移行'),
    # ─ Phase 4: テーパー期（W19-W23）────────────────
    19: (80,
         '5000m＋2000m @ 3:15/km（AM）',
         '1000m×6 @ 3:08/km（AM）',
         ''),
    20: (70,
         '3000m×2＋1000m×3 @ 3:15/km（AM）',
         '1000m×5 @ 3:08/km（AM）',
         ''),
    21: (62,
         '3000m＋2000m＋1000m @ 3:15/km（AM）',
         '1000m×4 @ 3:10/km（AM）',
         ''),
    22: (50,
         '2000m×2＋1000m×2（最終調整）',
         '800m×4 @ 3:03/km（最終シャープニング）',
         '最終調整週'),
    23: (38,
         '1000m×3 @ 3:15/km（軽め）',
         '800m×2＋400m×4（シャープニング）',
         '9/19選考の場合→W24が本番週'),
    24: (32,
         '1000m×2 @ 3:15/km（前日軽め）',
         '【選考会 9/19】10000m 32:30（理想32:00）',
         '★本番週【9/19候補】'),
    # 延長テーパー（9/26 または 10/3 選考の場合）
    25: (30,
         '1000m×3 @ 3:15/km（前日軽め）',
         '【選考会 9/26】10000m 32:30（理想32:00）',
         '★本番週【9/26候補】'),
    26: (28,
         '1000m×2 @ 3:15/km（前日軽め）',
         '【選考会 10/3】10000m 32:30（理想32:00）',
         '★本番週【10/3候補】'),
}

# ══════════════════════════════════════════════════════
#  スタイルヘルパー
# ══════════════════════════════════════════════════════
C_DARK   = '2B3A55'; C_MID = '4A6FA5'; C_WHITE = 'FFFFFF'
C_ALT    = 'F0F4FA'; C_GREEN = 'D4EDDA'; C_RED = 'FFADAD'; C_YELLOW = 'FFFDE7'
C_RACE   = 'FFD700'  # 選考会週のハイライト

def _bdr():
    sd = Side(style='thin', color='C0C0C0')
    return Border(left=sd, right=sd, top=sd, bottom=sd)

def cell(ws, r, c, v, bold=False, sz=10, bg=None, fg=C_DARK, h='center', wrap=False):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font      = Font(bold=bold, size=sz, color=fg, name='Yu Gothic')
    cl.alignment = Alignment(horizontal=h, vertical='center', wrap_text=wrap)
    cl.border    = _bdr()
    if bg:
        cl.fill = PatternFill('solid', fgColor=bg)
    return cl

def title_row(ws, r, text, ncols, bg=C_DARK):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font      = Font(bold=True, size=13, color=C_WHITE, name='Yu Gothic')
    c.fill      = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 26

def hdr_row(ws, r, headers, widths, bg=C_MID):
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell(ws, r, ci, h, bold=True, bg=bg, fg=C_WHITE, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 32

def fmd(dt):
    """Windows対応 月/日 フォーマット"""
    return f'{dt.month}/{dt.day}'

def week_start(wk_no):
    return SEASON_START + timedelta(weeks=wk_no - 1)

def phase_of(wk_no):
    for p in PHASES:
        if p['weeks'][0] <= wk_no <= p['weeks'][1]:
            return p
    return None

def zone_of(hr):
    if hr is None: return '—'
    if hr < MAX_HR * 0.60: return 'Zone1'
    if hr < MAX_HR * 0.70: return 'Zone2'
    if hr < MAX_HR * 0.80: return 'Zone3'
    if hr < MAX_HR * 0.90: return 'Zone4'
    return 'Zone5'

def pf(v):
    if v is None: return None
    try: return float(str(v).replace(',', ''))
    except: return None

def pdist(v):
    n = pf(v)
    if n is None or n == 0: return 0.0
    return round(n / 1000, 3) if n > 50 else round(n, 3)

def ptime_min(s):
    if s is None: return 0.0
    s = str(s)
    try:
        p = s.split(':')
        if len(p) == 3: return round(int(p[0]) * 60 + int(p[1]) + float(p[2]) / 60, 2)
        if len(p) == 2: return round(int(p[0]) + float(p[1]) / 60, 2)
    except:
        return 0.0
    return 0.0

# ══════════════════════════════════════════════════════
#  Excelファイル読み込み
# ══════════════════════════════════════════════════════
XLSX = 'hr_zone_analysis.xlsx'
PLAN_SHEET = 'フェーズ計画'
MGMT_SHEET = '週次管理'
RUN_TYPES  = {'ラン', 'トラックラン', 'トレッドミル', '屋内ラン'}

if not os.path.exists(XLSX):
    print(f'ERROR: {XLSX} が見つかりません。build_excel.py を先に実行してください。')
    input('Enterで終了...'); sys.exit(1)

print(f'\n読み込み中: {XLSX}')
wb = openpyxl.load_workbook(XLSX)

# ── Garminデータから週次実績を集計 ───────────────────
ws_g  = wb['Garminデータ']
hdr_g = [ws_g.cell(row=2, column=c).value for c in range(1, ws_g.max_column + 1)]

def find_col(hdr, *names):
    for name in names:
        for i, h in enumerate(hdr):
            if h and name in str(h): return i
    return None

CI_TYPE = find_col(hdr_g, 'アクティビティタイプ')
CI_DATE = find_col(hdr_g, '日付')
CI_DIST = find_col(hdr_g, '距離')
CI_TIME = find_col(hdr_g, 'タイム')
CI_AVHR = find_col(hdr_g, '平均心拍数')

weekly_actual = defaultdict(lambda: {
    'dist': 0.0, 'time': 0.0, 'sessions': 0,
    'z4_sessions': 0, 'hr_load': 0.0
})

for r in range(3, ws_g.max_row + 1):
    row = [ws_g.cell(row=r, column=c + 1).value for c in range(len(hdr_g))]
    if all(v is None for v in row): continue
    if CI_DATE is None or row[CI_DATE] is None: continue

    try:
        if isinstance(row[CI_DATE], datetime):
            dt = row[CI_DATE]
        else:
            dt = datetime.strptime(str(row[CI_DATE]), '%Y-%m-%d %H:%M:%S')
    except:
        continue

    if dt < SEASON_START: continue

    act_type = str(row[CI_TYPE] or '')
    dist     = pdist(row[CI_DIST]) if CI_DIST is not None else 0
    tmin     = ptime_min(row[CI_TIME]) if CI_TIME is not None else 0
    hr       = pf(row[CI_AVHR]) if CI_AVHR is not None else None

    delta = (dt - SEASON_START).days
    wk_no = delta // 7 + 1

    if act_type in RUN_TYPES:
        weekly_actual[wk_no]['dist'] += dist
    weekly_actual[wk_no]['sessions'] += 1
    if hr and zone_of(hr) in ('Zone4', 'Zone5'):
        weekly_actual[wk_no]['z4_sessions'] += 1
    if hr and tmin:
        weekly_actual[wk_no]['hr_load'] += round(hr * tmin)

print(f'  シーズン週次データ: {len(weekly_actual)} 週分集計完了')

# ── 既存メモを保持（週次管理シートのメモ列） ─────────
saved_notes = {}
if MGMT_SHEET in wb.sheetnames:
    ws_old = wb[MGMT_SHEET]
    for r in range(3, ws_old.max_row + 1):
        wk_val = ws_old.cell(row=r, column=1).value
        memo   = ws_old.cell(row=r, column=10).value
        if wk_val and memo:
            saved_notes[int(wk_val)] = str(memo)
    del wb[MGMT_SHEET]
    print(f'  既存メモ保持: {len(saved_notes)} 件')

if PLAN_SHEET in wb.sheetnames:
    del wb[PLAN_SHEET]

# ══════════════════════════════════════════════════════
#  Sheet A: フェーズ計画
# ══════════════════════════════════════════════════════
ws_plan = wb.create_sheet(PLAN_SHEET)

P_HDR = ['フェーズ', 'フェーズ名', '期間', '週数', '週距離目標(km)', '主要テーマ', '目標指標']
P_W   = [8, 16, 16, 6, 16, 38, 34]

title_row(ws_plan, 1,
    f'2026年シーズン トレーニング計画  ─  目標: 10000m {TARGET_TIME}  /  選考会 {RACE_DATE.strftime("%Y/%m/%d")}',
    len(P_HDR))
hdr_row(ws_plan, 2, P_HDR, P_W)

PHASE_THEMES = [
    'ダブルLT適応・怪我明け段階的増量 → 5/10東海大記録会 33:50',
    'ダブルLTのペース引き上げ・VO2max刺激導入・週100km突破',
    'レースペース（3:15/km）での走力確立・10000m完走力完成',
    '疲労抜き・スピードシャープニング・9月選考会本番',
]
PHASE_KPIS = [
    '東海大記録会 33:50 / ダブルLT定着（水・土各AM+PM）',
    '10000m換算 < 33:15 / 1000m×8本 @ 3:15/km',
    '中間TT 33:00 / 1000m×8本 @ 3:10/km',
    '選考会 32:30（理想32:00）',
]

for ri, p in enumerate(PHASES, 1):
    r   = ri + 2
    ws  = week_start(p['weeks'][0])
    we  = week_start(p['weeks'][1]) + timedelta(days=6)
    nwk = p['weeks'][1] - p['weeks'][0] + 1
    kms = [WEEKLY_PLAN[w][0] for w in range(p['weeks'][0], p['weeks'][1] + 1) if w in WEEKLY_PLAN]
    km_range = f'{min(kms)}〜{max(kms)}' if kms else '—'
    bg = p['color']; fg = p['fg']
    data = [
        f'Phase {p["id"]}', p['name'],
        f'{fmd(ws)}〜{fmd(we)}', nwk, km_range,
        PHASE_THEMES[ri - 1], PHASE_KPIS[ri - 1],
    ]
    for ci, v in enumerate(data, 1):
        h = 'left' if ci in (2, 6, 7) else 'center'
        cell(ws_plan, r, ci, v, bg=bg, fg=fg, h=h)

# ── パフォーマンス推移基準表 ─────────────────────────
ws_plan.row_dimensions[7].height = 8
title_row(ws_plan, 8, 'パフォーマンス基準（チェックポイント）', len(P_HDR), bg=C_MID)
PP_HDR = ['時期', '10000m換算', 'ペース目標', '1000mインターバル基準', '5000m換算']
PP_W   = [16, 14, 14, 26, 14]
for i, (h, w) in enumerate(zip(PP_HDR, PP_W), 1):
    cell(ws_plan, 9, i, h, bold=True, bg=C_MID, fg=C_WHITE)
    ws_plan.column_dimensions[get_column_letter(i)].width = max(P_W[i-1] if i <= len(P_W) else 0, w)

PP_DATA = [
    ('4月現在（復帰直後）',        '35:00〜36:00（推定）', '3:30〜3:36', 'ダブルLT適応中',         '< 16:30'),
    ('東海大学記録会（5/10）',     '33:50 目標',           '3:23',       '1000m×6 @ 3:28/km',     '< 16:00'),
    ('Phase2終了（6月末）',        '< 33:15',              '< 3:20',     '1000m×8 @ 3:15/km',     '< 15:45'),
    ('中間TT（7月下旬）',          '< 33:00 実測',         '< 3:18',     '1000m×8 @ 3:10/km',     '< 15:35'),
    ('Phase3終了（8月初旬）',      '< 32:45',              '< 3:16.5',   '1000m×8 @ 3:05/km',     '< 15:30'),
    ('選考会本番（9月）',          '32:30〜32:00',         '3:12〜3:15', '1000m×8 @ 3:03/km',     '15:25〜15:40'),
]
for ri, row_d in enumerate(PP_DATA, 1):
    r   = ri + 9
    alt = C_ALT if ri % 2 == 1 else C_WHITE
    for ci, v in enumerate(row_d, 1):
        cell(ws_plan, r, ci, v, bg=alt)

ws_plan.freeze_panes = 'A3'

# ══════════════════════════════════════════════════════
#  Sheet B: 週次管理
# ══════════════════════════════════════════════════════
ws_mgmt = wb.create_sheet(MGMT_SHEET)

M_HDR = ['週', '期間', 'Ph', '目標距離\n(km)', '実走行\n距離(km)', '達成率\n(%)',
         '高強度\nセッション', '水曜ポイント練習', '土曜ポイント練習', 'メモ（手入力）']
M_W   = [5, 13, 4, 8, 8, 7, 8, 32, 32, 22]

title_row(ws_mgmt, 1,
    f'週次トレーニング管理  ─  目標: 10000m {TARGET_TIME}  /  選考会: {RACE_DATE.strftime("%Y/%m/%d")}',
    len(M_HDR))
hdr_row(ws_mgmt, 2, M_HDR, M_W)
ws_mgmt.freeze_panes = 'A3'

today = datetime.now()
total_weeks  = max(WEEKLY_PLAN.keys())
# 選考会日程ごとの本番週番号
race_week_no = RACE_WEEK.get(RACE_DATE, 24)

for wk_no in range(1, total_weeks + 1):
    r    = wk_no + 2
    ws   = week_start(wk_no)
    we   = ws + timedelta(days=6)
    p    = phase_of(wk_no)
    plan = WEEKLY_PLAN.get(wk_no, (0, '—', '—', ''))

    target_km = plan[0]
    wed_menu  = plan[1]
    sat_menu  = plan[2]
    def_note  = plan[3]

    actual    = weekly_actual.get(wk_no, {})
    actual_km = round(actual.get('dist', 0.0), 1)
    z4_sess   = actual.get('z4_sessions', 0)
    ach_pct      = round(actual_km / target_km * 100, 1) if (target_km > 0 and actual_km > 0) else None
    is_race_week = (wk_no == race_week_no)
    is_extension = (wk_no > race_week_no)

    if is_race_week:
        row_bg = 'FFF9C4'          # 本番週: 薄金色
    elif is_extension:
        row_bg = 'EEEEEE'          # 延長週（他候補）: グレー
    elif ws <= today <= we:
        row_bg = 'FFFDE7'          # 今週: 薄黄
    elif we < today:
        row_bg = C_WHITE if wk_no % 2 == 0 else C_ALT  # 過去
    else:
        row_bg = 'F5F5F5'          # 未来

    # 達成率セルの色
    if ach_pct is None:
        ach_bg = row_bg
    elif ach_pct >= 95:
        ach_bg = C_GREEN
    elif ach_pct >= 80:
        ach_bg = C_YELLOW
    else:
        ach_bg = C_RED

    # フェーズ列の色
    ph_bg = p['color'] if p else C_WHITE
    ph_fg = p['fg']    if p else C_DARK
    ph_id = p['id']    if p else ''

    # メモ（手入力保持 → 初回はdefault文字）
    note_val = saved_notes.get(wk_no, def_note)

    data = [
        wk_no,
        f'{fmd(ws)}〜{fmd(we)}',
        ph_id,
        target_km,
        actual_km if actual_km > 0 else None,
        ach_pct,
        z4_sess if z4_sess > 0 else None,
        wed_menu,
        sat_menu,
        note_val,
    ]

    for ci, v in enumerate(data, 1):
        if ci == 3:
            cell(ws_mgmt, r, ci, v, bg=ph_bg, fg=ph_fg, bold=(is_race_week))
        elif ci == 6:
            cell(ws_mgmt, r, ci, v, bg=ach_bg, bold=(ach_pct is not None and ach_pct < 80))
        elif ci in (8, 9, 10):
            cell(ws_mgmt, r, ci, v, bg=row_bg, h='left', wrap=True)
        else:
            cell(ws_mgmt, r, ci, v, bg=row_bg)

    ws_mgmt.row_dimensions[r].height = 34

# 各選考会候補週を太枠で強調
for candidate_wk in [24, 25, 26]:
    if candidate_wk > total_weeks: continue
    for ci in range(1, len(M_HDR) + 1):
        cl = ws_mgmt.cell(row=candidate_wk + 2, column=ci)
        color = 'B8860B' if candidate_wk == race_week_no else '999999'
        thickness = 'medium' if candidate_wk == race_week_no else 'thin'
        sd = Side(style=thickness, color=color)
        cl.border = Border(left=sd, right=sd, top=sd, bottom=sd)
        cl.font   = Font(bold=(candidate_wk == race_week_no), size=10, color=C_DARK, name='Yu Gothic')

# ══════════════════════════════════════════════════════
#  シート順序の整理
# ══════════════════════════════════════════════════════
order = [
    'Garminデータ', 'ログ', '週間集計', '月別集計', 'ゾーン集計', 'ゾーン定義',
    PLAN_SHEET, MGMT_SHEET,
    'ポイント練習一覧', 'ラップデータ',
]
for i, name in enumerate(order):
    if name in wb.sheetnames:
        cur = wb.index(wb[name])
        wb.move_sheet(name, offset=i - cur)

wb.save(XLSX)
print(f'✓ 保存完了: {XLSX}')
print(f'  フェーズ計画・週次管理シートを更新しました（メモ保持: {len(saved_notes)} 件）')
