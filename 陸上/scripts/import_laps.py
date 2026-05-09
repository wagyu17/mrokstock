"""
import_laps.py  ─ ポイント練習ラップデータ インポートスクリプト
────────────────────────────────────────────────────────
使い方:
  1. GarminからダウンロードしたラップCSV(activity_*.csv)を
     「更新用トレーニングログ」フォルダに保存する
  2. 更新.bat を実行（update_excel.py の後に自動実行される）

  または単独で実行:
  > python import_laps.py

仕組み:
  - activity_*.csv の概要行（距離・タイム）を Activities*.csv と照合して日付を特定
  - 「ラップデータ」シートに1行/ラップで追記（重複は自動スキップ）
  - 「ポイント練習一覧」シートにセッションサマリーを追記
────────────────────────────────────────────────────────
"""
import sys, os, csv, glob
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════
#  定数
# ══════════════════════════════════════════════════════
XLSX       = 'hr_zone_analysis.xlsx'
LAP_DIR    = '更新用トレーニングログ'
LAP_SHEET  = 'ラップデータ'
LIST_SHEET = 'ポイント練習一覧'

C_DARK = '2B3A55'; C_MID = '4A6FA5'; C_WHITE = 'FFFFFF'; C_ALT = 'F0F4FA'

LAP_HEADERS = [
    '日付', '種別', 'ラップ', 'タイム', '累積時間', '距離(m)',
    'ペース(分/km)', '平均心拍', '最大心拍', '平均ピッチ(spm)',
    '接地時間(ms)', 'GCTバランス', '歩幅(m)', '上下動(cm)', '上下動比(%)',
]
LAP_WIDTHS = [13, 14, 6, 10, 10, 8, 13, 9, 9, 13, 12, 14, 9, 10, 12]

LIST_HEADERS = ['日付', '種別', '距離(m)', '合計タイム', '平均ペース', '平均心拍', '最大心拍', '平均ピッチ(spm)', 'メモ']
LIST_WIDTHS  = [13, 14, 10, 12, 13, 10, 10, 14, 30]

# ラップCSV列インデックス（0始まり）
I_LAP  = 0   # ラップ数
I_TIME = 1   # タイム
I_ATIME= 2   # 累積時間
I_DIST = 3   # 距離m
I_PACE = 4   # 平均ペース
I_AVHR = 6   # 平均心拍数
I_MXHR = 7   # 最大心拍数
I_PTCH = 14  # 平均ピッチ
I_GCT  = 15  # 接地時間ms
I_GCTB = 16  # GCTバランス%
I_STRD = 17  # 歩幅m
I_VOSC = 18  # 上下動cm
I_VRAT = 19  # 上下動比%

# ══════════════════════════════════════════════════════
#  スタイルヘルパー
# ══════════════════════════════════════════════════════
def _bdr():
    sd = Side(style='thin', color='C0C0C0')
    return Border(left=sd, right=sd, top=sd, bottom=sd)

def cell(ws, r, c, v, bold=False, sz=10, bg=None, fg=C_DARK, h='center'):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font      = Font(bold=bold, size=sz, color=fg, name='Yu Gothic')
    cl.alignment = Alignment(horizontal=h, vertical='center')
    cl.border    = _bdr()
    if bg:
        cl.fill = PatternFill('solid', fgColor=bg)
    return cl

def write_title_and_header(ws, title, headers, widths):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1, value=title)
    t.font      = Font(bold=True, size=13, color=C_WHITE, name='Yu Gothic')
    t.fill      = PatternFill('solid', fgColor=C_DARK)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell(ws, 2, ci, h, bold=True, bg=C_MID, fg=C_WHITE)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A3'

# ══════════════════════════════════════════════════════
#  パース関数
# ══════════════════════════════════════════════════════
def pf(v):
    if v is None or str(v).strip() in ('', '--'): return None
    try: return float(str(v).replace(',', ''))
    except: return None

def parse_time_to_sec(s):
    """タイム文字列を秒に変換 ('37:08' → 2228, '1:33.3' → 93.3)"""
    if not s or s == '--': return None
    s = str(s).strip()
    try:
        p = s.split(':')
        if len(p) == 3:
            return int(p[0]) * 3600 + int(p[1]) * 60 + float(p[2])
        if len(p) == 2:
            return int(p[0]) * 60 + float(p[1])
        return float(s)
    except:
        return None

# ══════════════════════════════════════════════════════
#  Activities CSV 読み込み
# ══════════════════════════════════════════════════════
def load_activities(lap_dir):
    """Activities*.csv からセッション一覧を読み込む。日付をキーとした辞書を返す。"""
    sessions = []
    for fpath in glob.glob(os.path.join(lap_dir, 'Activities*.csv')):
        with open(fpath, 'r', encoding='utf-8-sig') as fp:
            reader = csv.DictReader(fp)
            for row in reader:
                try:
                    dt_str   = row.get('日付', '')
                    dt       = datetime.strptime(dt_str, '%Y-%m-%d %H:%M:%S')
                    time_sec = parse_time_to_sec(row.get('タイム', ''))
                    sessions.append({
                        'date_str': dt.strftime('%Y/%m/%d'),
                        'type'    : row.get('アクティビティタイプ', ''),
                        'time_sec': time_sec,
                        'dist_raw': str(row.get('距離', '')).replace(',', ''),
                    })
                except:
                    continue
    return sessions

# ══════════════════════════════════════════════════════
#  日付照合（ラップCSV概要行 ↔ Activities）
# ══════════════════════════════════════════════════════
def find_session(summary_row, activities):
    """
    概要行のタイム（誤差30秒以内）で照合し、日付・種別を返す。
    距離もサブ条件として使う。
    """
    summary_sec  = parse_time_to_sec(summary_row[I_TIME])
    summary_dist = str(summary_row[I_DIST]).replace(',', '')

    best, best_diff = None, 999999
    for act in activities:
        if act['time_sec'] is None or summary_sec is None:
            continue
        diff = abs(act['time_sec'] - summary_sec)
        # 距離も比較（数値変換できる場合のみ）
        try:
            if act['dist_raw'] and summary_dist:
                if abs(float(act['dist_raw']) - float(summary_dist)) > 50:
                    continue  # 距離が大きくズレていたら除外
        except:
            pass
        if diff < best_diff:
            best_diff = diff
            best = act

    if best and best_diff <= 30:
        return best['date_str'], best['type']
    return None, None

# ══════════════════════════════════════════════════════
#  メイン処理
# ══════════════════════════════════════════════════════
if not os.path.exists(XLSX):
    print(f'ERROR: {XLSX} が見つかりません。build_excel.py を先に実行してください。')
    input('Enterで終了...')
    sys.exit(1)

print(f'\n読み込み中: {XLSX}')
wb = openpyxl.load_workbook(XLSX)

# ── シートの準備 ─────────────────────────────────────
for sname in [LIST_SHEET, LAP_SHEET]:
    if sname not in wb.sheetnames:
        wb.create_sheet(sname)
        print(f'  新規作成: 「{sname}」シート')

ws_lap  = wb[LAP_SHEET]
ws_list = wb[LIST_SHEET]

# 初回のみヘッダー書き込み
if ws_lap.cell(row=2, column=1).value != '日付':
    write_title_and_header(ws_lap, 'ポイント練習 ラップデータ（自動生成）', LAP_HEADERS, LAP_WIDTHS)

if ws_list.cell(row=2, column=1).value != '日付':
    write_title_and_header(ws_list, 'ポイント練習一覧（自動生成）', LIST_HEADERS, LIST_WIDTHS)

# ── 既存データ収集（重複防止） ────────────────────────
existing_laps = set()
for r in range(3, ws_lap.max_row + 1):
    d   = ws_lap.cell(row=r, column=1).value
    lap = ws_lap.cell(row=r, column=3).value
    if d and lap:
        existing_laps.add((str(d), str(lap)))

existing_sessions = set()
for r in range(3, ws_list.max_row + 1):
    d    = ws_list.cell(row=r, column=1).value
    dist = ws_list.cell(row=r, column=3).value
    if d and dist:
        existing_sessions.add((str(d), str(dist).replace(',', '')))

# ── Activitiesデータ読み込み ─────────────────────────
activities = load_activities(LAP_DIR)
print(f'  Activities: {len(activities)} セッション読み込み')

# ── ラップCSVを処理 ───────────────────────────────────
lap_files   = sorted(glob.glob(os.path.join(LAP_DIR, 'activity_*.csv')))
added_total = 0

if not lap_files:
    print(f'\n  ラップCSV(activity_*.csv)が「{LAP_DIR}」フォルダに見つかりません。')
else:
    for fpath in lap_files:
        fname = os.path.basename(fpath)
        print(f'\n処理中: {fname}')

        # CSV読み込み
        with open(fpath, 'r', encoding='utf-8-sig') as fp:
            reader = csv.reader(fp)
            _header = next(reader)   # ヘッダー行（読み飛ばし）
            all_rows = list(reader)

        # 概要行とラップ行を分離
        summary_row = None
        lap_rows    = []
        for row in all_rows:
            if not row: continue
            if row[I_LAP] == '概要':
                summary_row = row
            elif row[I_LAP].isdigit():
                lap_rows.append(row)

        if not summary_row:
            print(f'  ⚠ 概要行が見つかりません（インターバル形式など非対応フォーマット）。スキップします。')
            continue

        # 日付照合
        date_str, act_type = find_session(summary_row, activities)
        if not date_str:
            print(f'  ⚠ Activities.csv と照合できませんでした（タイム: {summary_row[I_TIME]}）。')
            print( '    ヒント: 最新の Activities*.csv を「更新用トレーニングログ」に置いてください。')
            continue

        print(f'  → {date_str}  {act_type}  {len(lap_rows)} ラップ')

        # ── ラップデータシートへ追記 ──────────────────
        added = 0
        r = max(ws_lap.max_row + 1, 3)
        for row in lap_rows:
            lap_no = row[I_LAP]
            key    = (date_str, lap_no)
            if key in existing_laps:
                continue

            alt = C_ALT if r % 2 == 1 else C_WHITE
            data = [
                date_str,
                act_type,
                int(lap_no),
                row[I_TIME],
                row[I_ATIME],
                row[I_DIST],
                row[I_PACE] if row[I_PACE] != '--' else None,
                pf(row[I_AVHR]),
                pf(row[I_MXHR]),
                pf(row[I_PTCH]),
                pf(row[I_GCT]),
                row[I_GCTB],
                pf(row[I_STRD]),
                pf(row[I_VOSC]),
                row[I_VRAT],
            ]
            for ci, v in enumerate(data, 1):
                h = 'left' if ci in (1, 2, 7, 12, 15) else 'center'
                cell(ws_lap, r, ci, v, bg=alt, h=h)

            existing_laps.add(key)
            r += 1
            added += 1

        # ── ポイント練習一覧シートへ追記 ─────────────
        dist_key = (date_str, str(summary_row[I_DIST]).replace(',', ''))
        if dist_key not in existing_sessions:
            lr  = max(ws_list.max_row + 1, 3)
            alt = C_ALT if lr % 2 == 1 else C_WHITE
            list_data = [
                date_str,
                act_type,
                summary_row[I_DIST],
                summary_row[I_TIME],
                summary_row[I_PACE] if summary_row[I_PACE] != '--' else None,
                pf(summary_row[I_AVHR]),
                pf(summary_row[I_MXHR]),
                pf(summary_row[I_PTCH]),
                '',   # メモ（手動記入欄）
            ]
            for ci, v in enumerate(list_data, 1):
                h = 'left' if ci in (1, 2, 5, 9) else 'center'
                cell(ws_list, lr, ci, v, bg=alt, h=h)
            existing_sessions.add(dist_key)

        print(f'  → {added} ラップ追加  (重複スキップ: {len(lap_rows) - added})')
        added_total += added

# ── シートを末尾に配置 ────────────────────────────────
order_end = [LIST_SHEET, LAP_SHEET]
for sname in order_end:
    if sname in wb.sheetnames:
        target_pos = len(wb.sheetnames) - 1
        current    = wb.index(wb[sname])
        wb.move_sheet(sname, offset=target_pos - current)

wb.save(XLSX)
print(f'\n✓ 保存完了: {XLSX}')
print(f'  合計 {added_total} ラップを追加しました')
