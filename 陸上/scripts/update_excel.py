"""
update_excel.py  ─ トレーニングログ 更新スクリプト
────────────────────────────────────────────────────────
使い方:
  1. Garmin Connect でCSVをダウンロードする
  2. hr_zone_analysis.xlsx を開き「Garminデータ」シートを選択
  3. CSVの全行（ヘッダー行含む）を A2 セルに貼り付ける
  4. Excel を保存して閉じる
  5. 更新.bat をダブルクリック
  6. Excel を開き直す → 全シート更新済み
────────────────────────────────────────────────────────
指標:
  HR負荷   = 平均心拍 × 運動時間(分)          ← メイン指標
  ゾーン負荷 = ゾーン加重(Z1=1〜Z5=5) × 時間(分) ← 強度の質を反映
  走行距離  = km（参考指標）
────────────────────────────────────────────────────────
"""
import sys, os
from datetime import datetime, timedelta
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import openpyxl.chart.label as chartlabel

# ══════════════════════════════════════════════════════
#  定数
# ══════════════════════════════════════════════════════
MAX_HR   = 187
ZONE_W   = {'Zone1':1,'Zone2':2,'Zone3':3,'Zone4':4,'Zone5':5,'—':1}
ZONE_BG  = {'Zone1':'AEF0D4','Zone2':'BDE0FB','Zone3':'FFF3B0',
            'Zone4':'FFCBA4','Zone5':'FFADAD','—':'E8E8E8'}
ZONE_FG  = {'Zone1':'2D6A4F','Zone2':'1D3557','Zone3':'7B5C00',
            'Zone4':'9B3B00','Zone5':'7D0000','—':'555555'}
ZCOLS_HEX = ['2D6A4F','1D3557','7B5C00','9B3B00','7D0000']
ZONE_NAMES = ['Zone1','Zone2','Zone3','Zone4','Zone5']
RUN_TYPES  = {'ラン','トラックラン','トレッドミル','屋内ラン'}

C_DARK='2B3A55'; C_MID='4A6FA5'; C_WHITE='FFFFFF'; C_ALT='F0F4FA'
DAYS_JP = ['月','火','水','木','金','土','日']
PIE_COLORS = ['4A6FA5','F28E2B','E15759','76B7B2','59A14F',
              'EDC948','B07AA1','FF9DA7','9C755F','BAB0AC']

# ══════════════════════════════════════════════════════
#  ファイル読み込み
# ══════════════════════════════════════════════════════
XLSX = 'hr_zone_analysis.xlsx'
if not os.path.exists(XLSX):
    print(f'ERROR: {XLSX} が見つかりません。build_excel.py を先に実行してください。')
    input('Enterで終了...'); sys.exit(1)

print(f'読み込み中: {XLSX}')
wb = openpyxl.load_workbook(XLSX)

if 'Garminデータ' not in wb.sheetnames:
    print('ERROR: 「Garminデータ」シートが見つかりません。')
    input('Enterで終了...'); sys.exit(1)

ws_g = wb['Garminデータ']
header = [ws_g.cell(row=2, column=c).value for c in range(1, ws_g.max_column+1)]

def find_col(hdr, *names):
    for name in names:
        for i,h in enumerate(hdr):
            if h and name in str(h): return i
    return None

CI_TYPE = find_col(header,'アクティビティタイプ')
CI_DATE = find_col(header,'日付')
CI_TITLE= find_col(header,'タイトル')
CI_DIST = find_col(header,'距離')
CI_CAL  = find_col(header,'カロリー')
CI_TIME = find_col(header,'タイム')
CI_AVHR = find_col(header,'平均心拍数')
CI_MXHR = find_col(header,'最大心拍数')
CI_PACE = find_col(header,'平均ペース')
CI_ELEV = find_col(header,'総上昇量')

if any(c is None for c in [CI_TYPE,CI_DATE,CI_AVHR,CI_TIME]):
    print('ERROR: Garminデータシートのヘッダーが認識できません。')
    print('       ヘッダー行を含めて A2 セルから貼り付けてください。')
    input('Enterで終了...'); sys.exit(1)

garmin_rows = []
for r in range(3, ws_g.max_row+1):
    row = [ws_g.cell(row=r, column=c+1).value for c in range(len(header))]
    if all(v is None for v in row): continue
    if row[CI_DATE] is None: continue
    garmin_rows.append(row)

print(f'  Garminデータ: {len(garmin_rows)} 行')

# ══════════════════════════════════════════════════════
#  スタイルヘルパー
# ══════════════════════════════════════════════════════
def fill(c): return PatternFill('solid', fgColor=c)
def bdr():
    s = Side(style='thin', color='C0C0C0')
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, r, c, v, bold=False, sz=10, bg=None, fg=C_DARK,
         h='center', wrap=False, bd=True, nf=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font      = Font(bold=bold, size=sz, color=fg, name='Yu Gothic')
    cl.alignment = Alignment(horizontal=h, vertical='center', wrap_text=wrap)
    if bg: cl.fill = PatternFill('solid', fgColor=bg)
    if bd: cl.border = bdr()
    if nf: cl.number_format = nf
    return cl

def title_row(ws, r, text, ncols, bg=C_DARK):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font  = Font(bold=True, size=13, color=C_WHITE, name='Yu Gothic')
    c.fill  = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 26

def hdr_row(ws, r, headers, bg=C_MID):
    for i,h in enumerate(headers,1):
        cell(ws, r, i, h, bold=True, bg=bg, fg=C_WHITE)
    ws.row_dimensions[r].height = 20

def set_col_widths(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════
#  パース関数
# ══════════════════════════════════════════════════════
def pf(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    s = str(v).strip()
    if s in ('--','','false','true'): return None
    try: return float(s.replace(',',''))
    except: return None

def pdist(v):
    n = pf(v)
    if n is None or n == 0: return 0.0
    return round(n/1000,3) if n > 50 else round(n,3)

def ptime(v):
    if v is None: return 0.0
    s = str(v).strip()
    if s in ('--',''): return 0.0
    try:
        p = s.split(':')
        if len(p)==3: return round(int(p[0])*60+int(p[1])+float(p[2])/60,2)
        if len(p)==2: return round(int(p[0])+float(p[1])/60,2)
    except: pass
    return 0.0

def ppace(v):
    if v in (None,'--',''): return None
    s = str(v).strip()
    if ':' not in s: return None
    p = s.split(':')
    try: return round(int(p[0])+int(p[1])/60,4)
    except: return None

def fmt_pace(v):
    if v is None: return '—'
    return f"{int(v)}:{int((v%1)*60):02d}"

def zone_of(hr):
    if hr is None: return '—'
    if hr < MAX_HR*0.60: return 'Zone1'
    if hr < MAX_HR*0.70: return 'Zone2'
    if hr < MAX_HR*0.80: return 'Zone3'
    if hr < MAX_HR*0.90: return 'Zone4'
    return 'Zone5'

def time_block(dt):
    h = dt.hour
    if 4 <= h < 12: return 'AM'
    if 12 <= h < 18: return 'PM'
    return 'Night'

def week_start(dt):
    return (dt - timedelta(days=dt.weekday())).strftime('%Y-%m-%d')

def parse_dt(v):
    if v is None: return None
    if isinstance(v, datetime): return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S','%Y/%m/%d %H:%M:%S','%Y-%m-%d'):
        try: return datetime.strptime(s[:19], fmt)
        except: pass
    return None

# ══════════════════════════════════════════════════════
#  セッションリスト構築
# ══════════════════════════════════════════════════════
def build_sessions(rows):
    sess = []; seen = set()
    for row in rows:
        dt = parse_dt(row[CI_DATE])
        if dt is None: continue
        key = (str(row[CI_DATE]), str(row[CI_TYPE]))
        if key in seen: continue
        seen.add(key)
        hr   = pf(row[CI_AVHR])
        mhr  = pf(row[CI_MXHR]) if CI_MXHR is not None else None
        dist = pdist(row[CI_DIST])
        cal  = pf(row[CI_CAL]) or 0
        tmin = ptime(row[CI_TIME])
        pace = ppace(row[CI_PACE]) if CI_PACE is not None else None
        elev = pf(row[CI_ELEV])  if CI_ELEV is not None else None
        typ  = str(row[CI_TYPE]) if row[CI_TYPE] else '不明'
        title= str(row[CI_TITLE]) if CI_TITLE is not None and row[CI_TITLE] else ''
        zn   = zone_of(hr)
        # ── メイン指標 ──
        hr_load   = round(hr * tmin) if hr and tmin else 0   # HR × 分
        zone_load = round(ZONE_W.get(zn,1) * tmin, 1)       # ゾーン加重 × 分
        sess.append({
            'type':typ, 'dt':dt, 'title':title,
            'dist':dist, 'cal':cal, 'time':tmin,
            'avg_hr':hr, 'max_hr':mhr,
            'pace':pace, 'zone':zn,
            'hr_load':hr_load, 'zone_load':zone_load,
            'week':week_start(dt), 'month':dt.strftime('%Y-%m'),
            'elev':elev,
        })
    return sorted(sess, key=lambda s: s['dt'], reverse=True)

sessions_all = build_sessions(garmin_rows)
sessions     = [s for s in sessions_all if s['type'] != '瞑想']
print(f'  有効セッション: {len(sessions_all)} 件（瞑想除く: {len(sessions)} 件）')

# ══════════════════════════════════════════════════════
#  集計関数
# ══════════════════════════════════════════════════════
def agg(rows):
    run = [s for s in rows if s['type'] in RUN_TYPES]
    zc  = defaultdict(int)
    zt  = defaultdict(float)   # ゾーン別時間
    zl  = defaultdict(float)   # ゾーン別HR負荷
    tc  = defaultdict(int)
    for s in rows:
        zn = s['zone']
        zc[zn]  += 1
        zt[zn]  += s['time']
        zl[zn]  += s['hr_load']
        tc[s['type']] += 1
    return {
        'n':         len(rows),
        'dist':      round(sum(s['dist']      for s in run),  1),
        'time':      round(sum(s['time']      for s in rows), 0),
        'hr_load':   int(sum(s['hr_load']     for s in rows)),
        'zone_load': round(sum(s['zone_load'] for s in rows), 1),
        'cal':       int(sum(s['cal']         for s in rows)),
        'zone_cnt':  dict(zc),
        'zone_time': dict(zt),
        'zone_load_d': dict(zl),
        'type':      dict(tc),
    }

all_weeks  = sorted(set(s['week']  for s in sessions))
all_months = sorted(set(s['month'] for s in sessions))
weekly  = {w: agg([s for s in sessions if s['week']==w])  for w in all_weeks}
monthly = {m: agg([s for s in sessions if s['month']==m]) for m in all_months}

# ══════════════════════════════════════════════════════
#  既存シートを再生成
# ══════════════════════════════════════════════════════
for name in ['ログ','週間集計','月別集計','ゾーン集計']:
    if name in wb.sheetnames: del wb[name]

# ══════════════════════════════════════════════════════
#  Sheet: ログ
# ══════════════════════════════════════════════════════
ws_log = wb.create_sheet('ログ')
LOG_H = ['No','日付','曜日','時間帯','種別',
         'HR負荷\n(HR×分)','ゾーン負荷\n(Z×分)',
         '走行距離\n(km)','タイム(分)',
         '平均心拍','最大心拍','ペース',
         'ゾーン','消費cal','上昇量(m)','RPE','メモ']
LOG_W = [5,13,5,8,16, 12,12, 10,10, 10,10,13, 8,10,10,6,30]

title_row(ws_log, 1, f'トレーニングログ（最終更新: {datetime.now().strftime("%Y/%m/%d %H:%M")}）', len(LOG_H))
hdr_row(ws_log, 2, LOG_H)
ws_log.freeze_panes = 'A3'
set_col_widths(ws_log, LOG_W)
ws_log.row_dimensions[2].height = 30

for ri, s in enumerate(sessions_all, 1):
    r   = ri + 2
    alt = C_ALT if ri%2==0 else C_WHITE
    zbg = ZONE_BG.get(s['zone'], C_WHITE)
    zfg = ZONE_FG.get(s['zone'], C_DARK)
    data = [
        ri,
        s['dt'].strftime('%Y/%m/%d'),
        DAYS_JP[s['dt'].weekday()],
        time_block(s['dt']),
        s['type'],
        s['hr_load']   if s['hr_load']   else None,
        s['zone_load'] if s['zone_load'] else None,
        round(s['dist'],2) if s['dist']  else None,
        round(s['time'],1) if s['time']  else None,
        int(s['avg_hr'])   if s['avg_hr'] else None,
        int(s['max_hr'])   if s['max_hr'] else None,
        fmt_pace(s['pace']),
        s['zone'],
        int(s['cal'])  if s['cal']  else None,
        int(s['elev']) if s['elev'] else None,
        '', '',
    ]
    for ci, v in enumerate(data, 1):
        bg = zbg if ci==13 else alt
        fg = zfg if ci==13 else C_DARK
        ah = 'left' if ci in (5,12,13,17) else 'center'
        cell(ws_log, r, ci, v, bg=bg, fg=fg, h=ah)

# ══════════════════════════════════════════════════════
#  Sheet: 週間集計
# ══════════════════════════════════════════════════════
ws_wk = wb.create_sheet('週間集計')
WK_H = ['週（月曜）','週末（日曜）','回数',
        'HR負荷\n(合計)','ゾーン負荷\n(合計)','走行距離\n(km)','時間(分)','消費cal',
        'Z1回','Z2回','Z3回','Z4回','Z5回',
        'Z1時間','Z2時間','Z3時間','Z4時間','Z5時間']
WK_W = [14,14,8, 12,12,12,10,10, 7,7,7,7,7, 8,8,8,8,8]

title_row(ws_wk, 1, f'週間トレーニング集計（{len(all_weeks)}週分）', len(WK_H))
hdr_row(ws_wk, 2, WK_H)
ws_wk.freeze_panes = 'A3'
set_col_widths(ws_wk, WK_W)
ws_wk.row_dimensions[2].height = 30

for ri, wk in enumerate(sorted(all_weeks, reverse=True), 1):
    r   = ri + 2
    alt = C_ALT if ri%2==0 else C_WHITE
    wk_dt = datetime.strptime(wk, '%Y-%m-%d')
    d  = weekly[wk]; zc = d['zone_cnt']; zt = d['zone_time']
    row_data = [
        wk_dt.strftime('%Y/%m/%d'),
        (wk_dt+timedelta(days=6)).strftime('%Y/%m/%d'),
        d['n'], d['hr_load'], d['zone_load'], d['dist'], d['time'], d['cal'],
        zc.get('Zone1',0), zc.get('Zone2',0), zc.get('Zone3',0), zc.get('Zone4',0), zc.get('Zone5',0),
        round(zt.get('Zone1',0),1), round(zt.get('Zone2',0),1),
        round(zt.get('Zone3',0),1), round(zt.get('Zone4',0),1), round(zt.get('Zone5',0),1),
    ]
    for ci, v in enumerate(row_data, 1):
        if 9<=ci<=13 and v:
            bg=ZONE_BG.get(f'Zone{ci-8}',alt); fg=ZONE_FG.get(f'Zone{ci-8}',C_DARK)
        elif 14<=ci<=18 and v:
            bg=ZONE_BG.get(f'Zone{ci-13}',alt); fg=ZONE_FG.get(f'Zone{ci-13}',C_DARK)
        else:
            bg=alt; fg=C_DARK
        cell(ws_wk, r, ci, v, bg=bg, fg=fg)

de = len(all_weeks) + 2

# 合計行
cell(ws_wk, de+1, 1, '合 計', bold=True, bg='CCCCCC')
for ci, key in [(3,'n'),(4,'hr_load'),(5,'zone_load'),(6,'dist'),(7,'time'),(8,'cal')]:
    v = sum(d[key] for d in weekly.values())
    cell(ws_wk, de+1, ci, round(v,1), bold=True, bg='CCCCCC')

# ── グラフ ──
c_ref = Reference(ws_wk, min_col=1, min_row=3, max_row=de)

def bar(title, col, color, anchor, h=13):
    ch = BarChart(); ch.type='col'; ch.style=10
    ch.title=title; ch.width=15; ch.height=h
    ch.add_data(Reference(ws_wk,min_col=col,min_row=2,max_row=de), titles_from_data=True)
    ch.set_categories(c_ref)
    ch.series[0].graphicalProperties.solidFill = color
    ws_wk.add_chart(ch, anchor)

# メイン2グラフ：HR負荷 & ゾーン負荷
bar('週別 HR負荷 (HR×分) ★メイン', 4, '2B3A55', 'T2')
bar('週別 ゾーン負荷 (Zone×分)',    5, '4A6FA5', 'T27')
# 参考グラフ：走行距離 & セッション数
bar('週別 走行距離 (km)',  6, '59A14F', 'T52')
bar('週別 セッション数',   3, 'E15759', 'T77')

# ゾーン時間 積み上げ棒
ch_zt = BarChart(); ch_zt.type='col'; ch_zt.grouping='stacked'; ch_zt.overlap=100
ch_zt.title='週別 ゾーン時間 (分)'; ch_zt.style=10; ch_zt.width=15; ch_zt.height=13
for zi, ci in enumerate(range(14,19)):
    ch_zt.add_data(Reference(ws_wk,min_col=ci,min_row=2,max_row=de), titles_from_data=True)
    ch_zt.series[zi].graphicalProperties.solidFill = ZCOLS_HEX[zi]
ch_zt.set_categories(c_ref)
ws_wk.add_chart(ch_zt, 'T102')

# 直近週 ゾーン負荷 円グラフ
latest_wk = sorted(all_weeks)[-1]
lw = weekly[latest_wk]
# 補助データ（AB列あたりに配置）
PIE_COL = len(WK_H) + 3   # = 21列目
cell(ws_wk, 2, PIE_COL,   '直近週ゾーン', bold=True, bg=C_MID, fg=C_WHITE)
cell(ws_wk, 2, PIE_COL+1, '回数',         bold=True, bg=C_MID, fg=C_WHITE)
cell(ws_wk, 2, PIE_COL+2, '時間(分)',     bold=True, bg=C_MID, fg=C_WHITE)
cell(ws_wk, 2, PIE_COL+3, 'HR負荷',       bold=True, bg=C_MID, fg=C_WHITE)
for zi, zn in enumerate(ZONE_NAMES):
    r = 3 + zi
    cell(ws_wk, r, PIE_COL,   zn,                           bg=ZONE_BG[zn], fg=ZONE_FG[zn])
    cell(ws_wk, r, PIE_COL+1, lw['zone_cnt'].get(zn,0),    bg=ZONE_BG[zn])
    cell(ws_wk, r, PIE_COL+2, round(lw['zone_time'].get(zn,0),1), bg=ZONE_BG[zn])
    cell(ws_wk, r, PIE_COL+3, int(lw['zone_load_d'].get(zn,0)),   bg=ZONE_BG[zn])

def make_pie(col_idx, title, anchor):
    pie = PieChart(); pie.title=title; pie.style=10
    pie.width=15; pie.height=13
    pie.add_data(Reference(ws_wk, min_col=col_idx, min_row=3, max_row=7))
    pie.set_categories(Reference(ws_wk, min_col=PIE_COL, min_row=3, max_row=7))
    pie.dataLabels = chartlabel.DataLabelList()
    pie.dataLabels.showPercent=True; pie.dataLabels.showCatName=True
    for idx,hx in enumerate(ZCOLS_HEX):
        pt=DataPoint(idx=idx); pt.graphicalProperties.solidFill=hx
        pie.series[0].dPt.append(pt)
    ws_wk.add_chart(pie, anchor)

make_pie(PIE_COL+1, f'直近週({latest_wk})\nゾーン回数',  'AJ2')
make_pie(PIE_COL+2, f'直近週({latest_wk})\nゾーン時間',  'AJ27')
make_pie(PIE_COL+3, f'直近週({latest_wk})\nHR負荷分布',  'AJ52')

# ══════════════════════════════════════════════════════
#  Sheet: 月別集計
# ══════════════════════════════════════════════════════
ws_mo = wb.create_sheet('月別集計')
MO_H = ['月','回数',
        'HR負荷\n(合計)','ゾーン負荷\n(合計)','走行距離\n(km)','時間(h)','消費cal',
        'Z1時間','Z2時間','Z3時間','Z4時間','Z5時間',
        'ラン','TM','バイク系','筋トレ','トラック']
MO_W = [10,8, 14,12,12,10,12, 8,8,8,8,8, 8,8,10,8,10]

title_row(ws_mo, 1, f'月別トレーニング集計（{len(all_months)}ヶ月）', len(MO_H))
hdr_row(ws_mo, 2, MO_H)
ws_mo.freeze_panes = 'A3'
set_col_widths(ws_mo, MO_W)
ws_mo.row_dimensions[2].height = 30

for ri, mo in enumerate(sorted(all_months, reverse=True), 1):
    r   = ri + 2
    alt = C_ALT if ri%2==0 else C_WHITE
    d   = monthly[mo]; zc = d['zone_cnt']; zt = d['zone_time']; tp = d['type']
    row_data = [
        mo, d['n'],
        d['hr_load'], d['zone_load'], d['dist'], round(d['time']/60,1), d['cal'],
        round(zt.get('Zone1',0),1), round(zt.get('Zone2',0),1),
        round(zt.get('Zone3',0),1), round(zt.get('Zone4',0),1), round(zt.get('Zone5',0),1),
        tp.get('ラン',0)+tp.get('屋内ラン',0), tp.get('トレッドミル',0),
        tp.get('屋内バイク',0)+tp.get('Xトレーナー',0)+tp.get('バイク',0),
        tp.get('筋力トレーニング',0), tp.get('トラックラン',0),
    ]
    for ci, v in enumerate(row_data, 1):
        if 8<=ci<=12 and v:
            bg=ZONE_BG.get(f'Zone{ci-7}',alt); fg=ZONE_FG.get(f'Zone{ci-7}',C_DARK)
        else:
            bg=alt; fg=C_DARK
        cell(ws_mo, r, ci, v, bg=bg, fg=fg)

me    = len(all_months) + 2
mc_ref = Reference(ws_mo, min_col=1, min_row=3, max_row=me)

# HR負荷推移（折れ線）★メイン
ch_hl = LineChart(); ch_hl.title='月別 HR負荷推移 ★メイン'; ch_hl.style=10
ch_hl.width=15; ch_hl.height=13; ch_hl.y_axis.title='HR×分'
ch_hl.add_data(Reference(ws_mo,min_col=3,min_row=2,max_row=me), titles_from_data=True)
ch_hl.set_categories(mc_ref)
ch_hl.series[0].graphicalProperties.line.solidFill = '2B3A55'
ch_hl.series[0].graphicalProperties.line.width = 25000
ws_mo.add_chart(ch_hl, 'S2')

# ゾーン負荷推移
ch_zl = LineChart(); ch_zl.title='月別 ゾーン負荷推移'; ch_zl.style=10
ch_zl.width=15; ch_zl.height=13; ch_zl.y_axis.title='Zone×分'
ch_zl.add_data(Reference(ws_mo,min_col=4,min_row=2,max_row=me), titles_from_data=True)
ch_zl.set_categories(mc_ref)
ch_zl.series[0].graphicalProperties.line.solidFill = '4A6FA5'
ch_zl.series[0].graphicalProperties.line.width = 25000
ws_mo.add_chart(ch_zl, 'S27')

# 走行距離推移（参考）
ch_ml = LineChart(); ch_ml.title='月別 走行距離推移 (km)'; ch_ml.style=10
ch_ml.width=15; ch_ml.height=13; ch_ml.y_axis.title='km'
ch_ml.add_data(Reference(ws_mo,min_col=5,min_row=2,max_row=me), titles_from_data=True)
ch_ml.set_categories(mc_ref)
ch_ml.series[0].graphicalProperties.line.solidFill = '59A14F'
ch_ml.series[0].graphicalProperties.line.width = 22000
ws_mo.add_chart(ch_ml, 'S52')

# ゾーン時間 積み上げ棒（月別）
ch_mz = BarChart(); ch_mz.type='col'; ch_mz.grouping='stacked'; ch_mz.overlap=100
ch_mz.title='月別 ゾーン時間 (分)'; ch_mz.style=10; ch_mz.width=15; ch_mz.height=13
for zi, ci in enumerate(range(8,13)):
    ch_mz.add_data(Reference(ws_mo,min_col=ci,min_row=2,max_row=me), titles_from_data=True)
    ch_mz.series[zi].graphicalProperties.solidFill = ZCOLS_HEX[zi]
ch_mz.set_categories(mc_ref)
ws_mo.add_chart(ch_mz, 'S77')

# 種別円グラフ
tot_types  = agg(sessions)['type']
type_items = [(k,v) for k,v in sorted(tot_types.items(), key=lambda x:-x[1])]
PIE_COL2 = len(MO_H) + 3
cell(ws_mo,2,PIE_COL2,'種別',bold=True,bg=C_MID,fg=C_WHITE)
cell(ws_mo,2,PIE_COL2+1,'回数',bold=True,bg=C_MID,fg=C_WHITE)
for i,(k,v) in enumerate(type_items):
    cell(ws_mo,3+i,PIE_COL2,k); cell(ws_mo,3+i,PIE_COL2+1,v)
pie_t = PieChart(); pie_t.title='全体 種別割合'; pie_t.style=10
pie_t.width=15; pie_t.height=13
pie_t.add_data(Reference(ws_mo,min_col=PIE_COL2+1,min_row=3,max_row=2+len(type_items)))
pie_t.set_categories(Reference(ws_mo,min_col=PIE_COL2,min_row=3,max_row=2+len(type_items)))
pie_t.dataLabels=chartlabel.DataLabelList()
pie_t.dataLabels.showPercent=True; pie_t.dataLabels.showCatName=True
for idx in range(len(type_items)):
    pt=DataPoint(idx=idx); pt.graphicalProperties.solidFill=PIE_COLORS[idx%len(PIE_COLORS)]
    pie_t.series[0].dPt.append(pt)
ws_mo.add_chart(pie_t, 'S102')

# ══════════════════════════════════════════════════════
#  Sheet: ゾーン集計
# ══════════════════════════════════════════════════════
ws_zn = wb.create_sheet('ゾーン集計')
ZN_H = ['ゾーン','心拍範囲','回数','割合(%)','合計時間(分)','HR負荷(合計)','ゾーン負荷(合計)','消費cal']
ZN_W = [16,16,8,10,14,16,16,12]

title_row(ws_zn, 1, f'心拍ゾーン別集計（最大心拍 {MAX_HR} bpm）', len(ZN_H))
hdr_row(ws_zn, 2, ZN_H)
for i,w in enumerate(ZN_W,1):
    ws_zn.column_dimensions[get_column_letter(i)].width = w

zone_all = defaultdict(lambda: {'n':0,'time':0,'hr_load':0,'zone_load':0,'cal':0})
for s in sessions:
    z = s['zone']
    zone_all[z]['n']         += 1
    zone_all[z]['time']      += s['time']
    zone_all[z]['hr_load']   += s['hr_load']
    zone_all[z]['zone_load'] += s['zone_load']
    zone_all[z]['cal']       += s['cal']

ZONES_DEF = [('Zone1',int(MAX_HR*0.50),int(MAX_HR*0.60)-1),
             ('Zone2',int(MAX_HR*0.60),int(MAX_HR*0.70)-1),
             ('Zone3',int(MAX_HR*0.70),int(MAX_HR*0.80)-1),
             ('Zone4',int(MAX_HR*0.80),int(MAX_HR*0.90)-1),
             ('Zone5',int(MAX_HR*0.90),MAX_HR),('—',None,None)]
total_n = len(sessions)

for ri,(zn,lo,hi) in enumerate(ZONES_DEF,1):
    r   = ri + 2
    rng = f'{lo}–{hi} bpm' if lo else '計測なし'
    bg  = ZONE_BG.get(zn,'E8E8E8'); fg = ZONE_FG.get(zn,C_DARK)
    d   = zone_all[zn]
    pct = round(d['n']/total_n*100,1) if total_n else 0
    for ci,v in enumerate([zn,rng,d['n'],pct,
                            round(d['time'],0),int(d['hr_load']),
                            round(d['zone_load'],1),int(d['cal'])],1):
        cell(ws_zn, r, ci, v, bg=bg, fg=fg)

tr = len(ZONES_DEF) + 3
cell(ws_zn,tr,1,'合 計',bold=True,bg='CCCCCC')
for ci,v in [(2,'—'),(3,total_n),(4,100.0),
             (5,round(sum(d['time']      for d in zone_all.values()),0)),
             (6,int(sum(d['hr_load']    for d in zone_all.values()))),
             (7,round(sum(d['zone_load']for d in zone_all.values()),1)),
             (8,int(sum(d['cal']        for d in zone_all.values())))]:
    cell(ws_zn,tr,ci,v,bold=True,bg='CCCCCC')

# HR負荷 円グラフ（ゾーン別）
PIE_Z_COL = len(ZN_H) + 3
for zi,zn in enumerate(ZONE_NAMES):
    r = 3 + zi
    cell(ws_zn, r, PIE_Z_COL,   zn,                               bg=ZONE_BG[zn], fg=ZONE_FG[zn])
    cell(ws_zn, r, PIE_Z_COL+1, int(zone_all[zn]['hr_load']),     bg=ZONE_BG[zn])
    cell(ws_zn, r, PIE_Z_COL+2, round(zone_all[zn]['time'],1),    bg=ZONE_BG[zn])

def pie_zone(col_idx, title, anchor):
    p = PieChart(); p.title=title; p.style=10; p.width=15; p.height=13
    p.add_data(Reference(ws_zn,min_col=col_idx,min_row=3,max_row=7))
    p.set_categories(Reference(ws_zn,min_col=PIE_Z_COL,min_row=3,max_row=7))
    p.dataLabels=chartlabel.DataLabelList()
    p.dataLabels.showPercent=True; p.dataLabels.showCatName=True
    for idx,hx in enumerate(ZCOLS_HEX):
        pt=DataPoint(idx=idx); pt.graphicalProperties.solidFill=hx
        p.series[0].dPt.append(pt)
    ws_zn.add_chart(p, anchor)

pie_zone(PIE_Z_COL+1, '全体 ゾーン別HR負荷', 'J2')
pie_zone(PIE_Z_COL+2, '全体 ゾーン別時間',   'J27')

# ══════════════════════════════════════════════════════
#  ゾーン定義（存在しなければ作成）
# ══════════════════════════════════════════════════════
if 'ゾーン定義' not in wb.sheetnames:
    ws_def = wb.create_sheet('ゾーン定義')
    DEF_H = ['ゾーン','下限(bpm)','上限(bpm)','加重','目的']
    DEF_W = [12,12,12,8,36]
    PURPOSES = ['積極的回復・疲労抜き','脂肪燃焼・有酸素基礎','有酸素能力向上','乳酸閾値トレーニング','最大酸素摂取量・スピード']
    PCT_LBL  = ['50–60 %','60–70 %','70–80 %','80–90 %','90–100 %']
    title_row(ws_def, 1, f'心拍数ゾーン定義（最大心拍数 {MAX_HR} bpm）', len(DEF_H))
    hdr_row(ws_def, 2, DEF_H)
    set_col_widths(ws_def, DEF_W)
    for ri,(zn,lo,hi) in enumerate(ZONES_DEF[:5],1):
        r=ri+2; bg=ZONE_BG[zn]; fg=ZONE_FG[zn]
        for ci,v in enumerate([zn,lo,hi,ZONE_W[zn],PURPOSES[ri-1]],1):
            cell(ws_def,r,ci,v,bg=bg,fg=fg,h='left' if ci==5 else 'center')

# ══════════════════════════════════════════════════════
#  シート順序
# ══════════════════════════════════════════════════════
order = ['Garminデータ','ログ','週間集計','月別集計','ゾーン集計','ゾーン定義']
for i, name in enumerate(order):
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=-wb.index(wb[name])+i)

# ══════════════════════════════════════════════════════
#  保存・サマリー表示
# ══════════════════════════════════════════════════════
wb.save(XLSX)
tot = agg(sessions)
print(f'\n✓ 更新完了: {XLSX}')
print(f'  期間        : {min(s["dt"] for s in sessions).strftime("%Y/%m/%d")} 〜 {max(s["dt"] for s in sessions).strftime("%Y/%m/%d")}')
print(f'  総セッション : {tot["n"]} 件')
print(f'  総HR負荷    : {tot["hr_load"]:,} HR×分')
print(f'  総ゾーン負荷 : {tot["zone_load"]:,} Zone×分')
print(f'  総走行距離   : {tot["dist"]:.1f} km')
print(f'  週間集計     : {len(all_weeks)} 週')
print(f'  月別集計     : {len(all_months)} ヶ月')
input('\nEnterキーで閉じる...')
