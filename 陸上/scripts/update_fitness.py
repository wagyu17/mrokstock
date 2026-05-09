"""
update_fitness.py  ─ フィットネス管理シート（CTL/ATL/TSB）更新スクリプト
────────────────────────────────────────────────────────
機能:
  1. 更新用トレーニングログ/ 内の Activities*.csv を読み込み、
     Garminデータシートに新規行を追記（重複は日付+種別で自動スキップ）
  2. Garminデータ全件からTSSを計算し、CTL/ATL/TSBを算出して
     「フィットネス」シートを生成/更新

指標説明:
  TSS  (Training Stress Score)  ─ 1セッションの負荷
         = (時間[h]) × 100 × ((avg_hr - REST_HR) / (THR_HR - REST_HR))²
  ATL  (Acute Training Load)    ─ 直近7日平均 ≒ 疲労
  CTL  (Chronic Training Load)  ─ 直近42日平均 ≒ フィットネス
  TSB  (Training Stress Balance)─ CTL − ATL ≒ コンディション

定数（要調整）:
  REST_HR    = 45   bpm（安静時心拍）
  THR_HR     = 150  bpm（LT2 / Z4境界 = 閾値心拍）
  MAX_HR     = 187  bpm
────────────────────────────────────────────────────────
"""
import sys, os, csv, glob, math
from datetime import datetime, date, timedelta
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ══════════════════════════════════════════════════════
#  定数
# ══════════════════════════════════════════════════════
XLSX        = 'hr_zone_analysis.xlsx'
GARMIN_SHEET= 'Garminデータ'
FIT_SHEET   = 'フィットネス'
ACT_DIR     = '更新用トレーニングログ'

REST_HR  = 45    # 安静時心拍
THR_HR   = 150   # 閾値心拍（LT2, Z4境界）
MAX_HR   = 187

ATL_TC   = 7.0   # ATL時定数（日）
CTL_TC   = 42.0  # CTL時定数（日）
SHOW_DAYS= 120   # シートに表示する日数

C_DARK   = '2B3A55'
C_MID    = '4A6FA5'
C_WHITE  = 'FFFFFF'
C_ALT    = 'F0F4FA'
C_GREEN  = 'D6EFD8'
C_YELLOW = 'FFF9C4'
C_RED    = 'FDDDE6'

RUN_TYPES = {'ラン', 'トラックラン', 'トレッドミル', '屋内ラン'}
# ランニング以外のクロストレーニング係数（負荷を何倍に換算するか）
CROSS_COEF = {'Xトレーナー': 0.7, '屋内バイク': 0.6, '筋力トレーニング': 0.3}

# ══════════════════════════════════════════════════════
#  ヘルパー
# ══════════════════════════════════════════════════════
def pf(v):
    if v is None: return None
    s = str(v).strip()
    if s in ('--', '', 'false', 'true'): return None
    try: return float(s.replace(',', ''))
    except: return None

def ptime(v):
    if v is None: return 0.0
    s = str(v).strip()
    if s in ('--', ''): return 0.0
    try:
        p = s.split(':')
        if len(p) == 3: return int(p[0]) * 60 + int(p[1]) + float(p[2]) / 60
        if len(p) == 2: return int(p[0]) + float(p[1]) / 60
    except: pass
    return 0.0

def parse_dt(v):
    if v is None: return None
    if isinstance(v, (datetime, date)):
        return v if isinstance(v, datetime) else datetime(v.year, v.month, v.day)
    s = str(v).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d'):
        try: return datetime.strptime(s[:19], fmt)
        except: pass
    return None

def fill(c): return PatternFill('solid', fgColor=c)
def bdr():
    s = Side(style='thin', color='C0C0C0')
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, r, c, v, bold=False, sz=10, bg=None, fg=C_DARK,
         h='center', bd=True, nf=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font      = Font(bold=bold, size=sz, color=fg, name='Yu Gothic')
    cl.alignment = Alignment(horizontal=h, vertical='center')
    if bg: cl.fill = PatternFill('solid', fgColor=bg)
    if bd: cl.border = bdr()
    if nf: cl.number_format = nf
    return cl

def title_row(ws, r, text, ncols):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font  = Font(bold=True, size=13, color=C_WHITE, name='Yu Gothic')
    c.fill  = PatternFill('solid', fgColor=C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 26

def hdr_row(ws, r, headers, bg=C_MID):
    for i, h in enumerate(headers, 1):
        cell(ws, r, i, h, bold=True, bg=bg, fg=C_WHITE)
    ws.row_dimensions[r].height = 20

# ══════════════════════════════════════════════════════
#  Excel 読み込み
# ══════════════════════════════════════════════════════
if not os.path.exists(XLSX):
    print(f'ERROR: {XLSX} が見つかりません。')
    input('Enterで終了...'); sys.exit(1)

print(f'読み込み中: {XLSX}')
wb = openpyxl.load_workbook(XLSX)

if GARMIN_SHEET not in wb.sheetnames:
    print(f'ERROR: 「{GARMIN_SHEET}」シートが見つかりません。')
    input('Enterで終了...'); sys.exit(1)

ws_g = wb[GARMIN_SHEET]

# ── ヘッダー列インデックスを特定 ──
header = [ws_g.cell(row=2, column=c).value for c in range(1, ws_g.max_column + 1)]

def find_col(hdr, *names):
    for name in names:
        for i, h in enumerate(hdr):
            if h and name in str(h): return i
    return None

CI_TYPE = find_col(header, 'アクティビティタイプ')
CI_DATE = find_col(header, '日付')

# ══════════════════════════════════════════════════════
#  Part 1: Activities*.csv → Garminデータ シートへ追記
# ══════════════════════════════════════════════════════
print('\n[Part 1] Activities CSVインポート')

# 既存の (日付文字列, 種別) セットを収集
existing_keys = set()
for r in range(3, ws_g.max_row + 1):
    d = ws_g.cell(row=r, column=CI_DATE + 1).value
    t = ws_g.cell(row=r, column=CI_TYPE + 1).value
    if d is not None:
        existing_keys.add((str(d), str(t) if t else ''))

# Activities*.csv を走査
csv_files = sorted(glob.glob(os.path.join(ACT_DIR, 'Activities*.csv')))
if not csv_files:
    print('  Activities*.csv が見つかりません。スキップ。')
else:
    total_added = 0
    for fpath in csv_files:
        print(f'  読み込み: {os.path.basename(fpath)}')
        try:
            with open(fpath, encoding='utf-8-sig', newline='') as f:
                reader = csv.reader(f)
                rows = list(reader)
        except Exception as e:
            print(f'  ERROR: {e}')
            continue

        if not rows:
            continue

        # 1行目がヘッダー
        csv_header = rows[0]
        added = 0

        # 列数を既存シートの列数に合わせて揃える
        n_cols = len(header)

        for row in rows[1:]:
            if len(row) < 2: continue
            key = (row[1], row[0])  # (日付, 種別)
            if key in existing_keys:
                continue
            existing_keys.add(key)

            # 行を既存ヘッダーの列順に並べる（余分な列は切り捨て、不足はNullで補完）
            padded = (row + [''] * n_cols)[:n_cols]
            next_row = ws_g.max_row + 1
            for ci, val in enumerate(padded, 1):
                ws_g.cell(row=next_row, column=ci, value=val if val not in ('--', '') else None)
            added += 1
            total_added += 1

        print(f'    追加: {added} 行')

    print(f'  合計追加: {total_added} 行')

# ══════════════════════════════════════════════════════
#  Part 2: Garminデータ全件からTSS計算 → CTL/ATL/TSB
# ══════════════════════════════════════════════════════
print('\n[Part 2] TSS / CTL / ATL / TSB 計算')

# 全セッションを収集
sessions = []
seen_keys = set()

for r in range(3, ws_g.max_row + 1):
    row = [ws_g.cell(row=r, column=c + 1).value for c in range(len(header))]
    if all(v is None for v in row): continue

    dt   = parse_dt(row[CI_DATE] if CI_DATE is not None else None)
    if dt is None: continue
    # 異常な年はスキップ
    if dt.year < 2020 or dt.year > 2030: continue

    typ  = str(row[CI_TYPE]) if CI_TYPE is not None and row[CI_TYPE] else ''
    if typ == '瞑想': continue

    key = (dt.strftime('%Y-%m-%d %H:%M:%S'), typ)
    if key in seen_keys: continue
    seen_keys.add(key)

    # タイム列インデックス（update_excel.pyと同じfind_col）
    ci_time = find_col(header, 'タイム')
    ci_avhr = find_col(header, '平均心拍数')
    tmin = ptime(row[ci_time]) if ci_time is not None else 0.0
    avhr = pf(row[ci_avhr])    if ci_avhr is not None else None

    sessions.append({'dt': dt, 'type': typ, 'tmin': tmin, 'avg_hr': avhr})

sessions.sort(key=lambda s: s['dt'])
print(f'  有効セッション: {len(sessions)} 件')

# ── TSS計算 ──
def calc_tss(avg_hr, tmin, typ):
    """HR-based TSS (閾値=THR_HR で 100TSS/h に正規化)"""
    if not avg_hr or not tmin or avg_hr <= REST_HR:
        return 0.0
    hr_ratio = (avg_hr - REST_HR) / (THR_HR - REST_HR)
    hr_ratio = min(hr_ratio, 2.0)   # 上限2倍でキャップ
    tss = (tmin / 60.0) * 100.0 * (hr_ratio ** 2)
    # クロストレーニング係数
    coef = CROSS_COEF.get(typ, 1.0) if typ not in RUN_TYPES else 1.0
    return round(tss * coef, 1)

for s in sessions:
    s['tss'] = calc_tss(s['avg_hr'], s['tmin'], s['type'])

# ── 日次TSS集計（同日複数セッションを合計）──
daily_tss = defaultdict(float)
for s in sessions:
    d = s['dt'].date()
    daily_tss[d] += s['tss']

# ── CTL/ATL をday-by-dayで計算 ──
if not daily_tss:
    print('  セッションデータがありません。スキップ。')
    sys.exit(0)

start_date = min(daily_tss.keys())
end_date   = date.today()

atl_k = 1.0 - math.exp(-1.0 / ATL_TC)
ctl_k = 1.0 - math.exp(-1.0 / CTL_TC)

atl = 0.0
ctl = 0.0
fitness_records = []   # (date, tss, ctl, atl, tsb)

d = start_date
while d <= end_date:
    tss = daily_tss.get(d, 0.0)
    atl = atl * (1 - atl_k) + tss * atl_k
    ctl = ctl * (1 - ctl_k) + tss * ctl_k
    tsb = ctl - atl
    fitness_records.append((d, round(tss, 1), round(ctl, 1), round(atl, 1), round(tsb, 1)))
    d += timedelta(days=1)

print(f'  計算期間: {start_date} 〜 {end_date}')
today_rec = fitness_records[-1]
print(f'  本日: TSS={today_rec[1]:.0f}  CTL={today_rec[2]:.1f}  ATL={today_rec[3]:.1f}  TSB={today_rec[4]:.1f}')

# ══════════════════════════════════════════════════════
#  Part 3: フィットネスシート生成
# ══════════════════════════════════════════════════════
print('\n[Part 3] フィットネスシート生成')

if FIT_SHEET in wb.sheetnames:
    del wb[FIT_SHEET]

# ログシートの後ろに挿入
try:
    log_idx = wb.sheetnames.index('ログ')
    ws_f = wb.create_sheet(FIT_SHEET, log_idx + 1)
except ValueError:
    ws_f = wb.create_sheet(FIT_SHEET)

# ── レイアウト定数 ──
NCOLS = 7
FIT_H = ['日付', 'TSS\n(本日負荷)', 'CTL\n(42日フィットネス)', 'ATL\n(7日疲労)',
         'TSB\n(コンディション)', '判定', 'メモ']
FIT_W = [13, 13, 18, 13, 18, 12, 25]

title_row(ws_f, 1, f'フィットネス管理  CTL / ATL / TSB（更新: {datetime.now().strftime("%Y/%m/%d %H:%M")}）', NCOLS)
hdr_row(ws_f, 2, FIT_H)
ws_f.freeze_panes = 'A3'
ws_f.row_dimensions[2].height = 30

for i, w in enumerate(FIT_W, 1):
    ws_f.column_dimensions[get_column_letter(i)].width = w

# ── データ行（直近SHOW_DAYS日、新しい順）──
recent = fitness_records[-SHOW_DAYS:]
recent_reversed = list(reversed(recent))

def tsb_label(tsb):
    if tsb >= 10:  return '◎ ピーク'
    if tsb >= 0:   return '○ 良好'
    if tsb >= -15: return '△ 普通'
    if tsb >= -30: return '▲ 要注意'
    return           '✕ 疲労蓄積'

def tsb_bg(tsb):
    if tsb >= 10:  return C_GREEN
    if tsb >= 0:   return 'E8F5E9'
    if tsb >= -15: return C_WHITE
    if tsb >= -30: return C_YELLOW
    return           C_RED

for ri, (d, tss, ctl, atl, tsb) in enumerate(recent_reversed, 1):
    r   = ri + 2
    alt = C_ALT if ri % 2 == 0 else C_WHITE
    bg  = tsb_bg(tsb)
    is_today = (d == date.today())
    row_bg = '[ FFF9C4' if is_today else alt

    cell(ws_f, r, 1, d.strftime('%Y-%m-%d (%a)'), bg=('FFFDE7' if is_today else alt), h='left')
    cell(ws_f, r, 2, tss,  bg=alt, nf='0.0')
    cell(ws_f, r, 3, ctl,  bg=alt, nf='0.0')
    cell(ws_f, r, 4, atl,  bg=alt, nf='0.0')
    cell(ws_f, r, 5, tsb,  bg=bg,  nf='0.0',
         bold=(tsb >= 10 or tsb <= -30),
         fg=('7D0000' if tsb <= -30 else ('2D6A4F' if tsb >= 10 else C_DARK)))
    cell(ws_f, r, 6, tsb_label(tsb), bg=bg,
         fg=('7D0000' if tsb <= -30 else ('2D6A4F' if tsb >= 10 else C_DARK)))
    cell(ws_f, r, 7, '', bg=alt, h='left')

# ── サマリーボックス（右上） ──
SUMMARY_COL = 9
ws_f.column_dimensions[get_column_letter(SUMMARY_COL)].width = 16
ws_f.column_dimensions[get_column_letter(SUMMARY_COL + 1)].width = 10

td = fitness_records[-1]
summary_data = [
    ('本日のTSS',   f'{td[1]:.0f}'),
    ('CTL (フィットネス)', f'{td[2]:.1f}'),
    ('ATL (疲労)',  f'{td[3]:.1f}'),
    ('TSB (コンディション)', f'{td[4]:.1f}'),
    ('判定',        tsb_label(td[4])),
]

ws_f.merge_cells(start_row=1, start_column=SUMMARY_COL, end_row=1, end_column=SUMMARY_COL + 1)
sc = ws_f.cell(row=1, column=SUMMARY_COL, value='本日のサマリー')
sc.font  = Font(bold=True, size=11, color=C_WHITE, name='Yu Gothic')
sc.fill  = PatternFill('solid', fgColor=C_MID)
sc.alignment = Alignment(horizontal='center', vertical='center')

for si, (label, val) in enumerate(summary_data, 2):
    lc = ws_f.cell(row=si, column=SUMMARY_COL, value=label)
    lc.font = Font(size=10, name='Yu Gothic', color=C_DARK)
    lc.fill = PatternFill('solid', fgColor=C_ALT)
    lc.border = bdr()
    lc.alignment = Alignment(horizontal='left', vertical='center')

    vc = ws_f.cell(row=si, column=SUMMARY_COL + 1, value=val)
    vc.font = Font(size=10, bold=True, name='Yu Gothic',
                   color=('7D0000' if td[4] <= -30 else ('2D6A4F' if td[4] >= 10 else C_DARK)))
    vc.fill = PatternFill('solid', fgColor=tsb_bg(td[4]) if label == '判定' else C_ALT)
    vc.border = bdr()
    vc.alignment = Alignment(horizontal='center', vertical='center')

# ── ラインチャート（CTL/ATL/TSB）──
chart_data_start = len(recent_reversed) + 3   # データの直後
chart_start_row  = 8   # チャートを配置する行（右側）

# チャート用データは古い順でないといけないのでrecentを使用
CHART_ROWS = min(90, len(recent))
chart_recent = recent[-CHART_ROWS:]

# 隠し列にチャート用データを書き込む（列L以降）
CHART_COL_BASE = 12  # L列
ws_f.cell(row=1, column=CHART_COL_BASE,     value='_date')
ws_f.cell(row=1, column=CHART_COL_BASE + 1, value='CTL')
ws_f.cell(row=1, column=CHART_COL_BASE + 2, value='ATL')
ws_f.cell(row=1, column=CHART_COL_BASE + 3, value='TSB')

for ci, (d, tss, ctl, atl, tsb) in enumerate(chart_recent, 2):
    ws_f.cell(row=ci, column=CHART_COL_BASE,     value=d.strftime('%m/%d'))
    ws_f.cell(row=ci, column=CHART_COL_BASE + 1, value=ctl)
    ws_f.cell(row=ci, column=CHART_COL_BASE + 2, value=atl)
    ws_f.cell(row=ci, column=CHART_COL_BASE + 3, value=tsb)

chart = LineChart()
chart.title    = 'CTL / ATL / TSB（直近90日）'
chart.style    = 10
chart.y_axis.title = '負荷'
chart.x_axis.title = '日付'
chart.width  = 22
chart.height = 14

n = len(chart_recent)
ctl_ref = Reference(ws_f, min_col=CHART_COL_BASE + 1, min_row=1, max_row=n + 1)
atl_ref = Reference(ws_f, min_col=CHART_COL_BASE + 2, min_row=1, max_row=n + 1)
tsb_ref = Reference(ws_f, min_col=CHART_COL_BASE + 3, min_row=1, max_row=n + 1)

chart.add_data(ctl_ref, titles_from_data=True)
chart.add_data(atl_ref, titles_from_data=True)
chart.add_data(tsb_ref, titles_from_data=True)

# CTL=青, ATL=橙, TSB=緑
from openpyxl.drawing.fill import ColorChoice
from openpyxl.chart.data_source import NumDataSource, NumRef
for idx, (color, width) in enumerate(
        [('4A6FA5', 25000), ('F28E2B', 20000), ('4CAF50', 15000)]):
    s = chart.series[idx]
    s.graphicalProperties.line.solidFill = color
    s.graphicalProperties.line.width     = width
    s.smooth = True

# TSBは0ラインを参考にするため2軸にする
chart.series[2].axId = 200

ws_f.add_chart(chart, f'{get_column_letter(SUMMARY_COL)}8')

print(f'  「{FIT_SHEET}」シート生成完了')

# ══════════════════════════════════════════════════════
#  保存
# ══════════════════════════════════════════════════════
print(f'\n保存中: {XLSX}')
wb.save(XLSX)
print('完了。')
print(f'\n現在のフィットネス状態:')
print(f'  CTL (フィットネス) = {today_rec[2]:.1f}')
print(f'  ATL (疲労)         = {today_rec[3]:.1f}')
print(f'  TSB (コンディション) = {today_rec[4]:.1f}  → {tsb_label(today_rec[4])}')
