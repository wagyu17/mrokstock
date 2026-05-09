"""
build_excel.py
初回実行用: training_log_2026.csv から Garminデータシートを生成し、
全分析シートを構築する。
以降は update_excel.py を使う。
"""
import csv, sys
from datetime import datetime, timedelta
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import openpyxl.chart.label as chartlabel

# ──────────────────────────────────────────────────────
#  CSV 読み込み
# ──────────────────────────────────────────────────────
with open('training_log_2026.csv', 'r', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    garmin_header = next(reader)
    garmin_rows   = list(reader)

print(f"読み込み: {len(garmin_rows)} 件")

wb = openpyxl.Workbook()

# ──────────────────────────────────────────────────────
#  共通スタイル
# ──────────────────────────────────────────────────────
C_DARK   = '2B3A55'
C_MID    = '4A6FA5'
C_WHITE  = 'FFFFFF'
C_ALT    = 'F0F4FA'
C_WARN   = 'FFF3CD'
C_GREEN  = 'D4EDDA'
ZONE_BG  = {'Zone1':'AEF0D4','Zone2':'BDE0FB','Zone3':'FFF3B0',
            'Zone4':'FFCBA4','Zone5':'FFADAD','—':'E8E8E8'}
ZONE_FG  = {'Zone1':'2D6A4F','Zone2':'1D3557','Zone3':'7B5C00',
            'Zone4':'9B3B00','Zone5':'7D0000','—':'555555'}

def sty(bold=False, sz=10, fg=C_DARK, name='Yu Gothic'):
    return Font(bold=bold, size=sz, color=fg, name=name)

def fill(c): return PatternFill('solid', fgColor=c)

def bdr(c='C0C0C0', s='thin'):
    sd = Side(style=s, color=c)
    return Border(left=sd, right=sd, top=sd, bottom=sd)

def aln(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cell(ws, r, c, v, bold=False, sz=10, bg=None, fg=C_DARK,
         h='center', wrap=False, bd=True, nf=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = Font(bold=bold, size=sz, color=fg, name='Yu Gothic')
    cl.alignment = Alignment(horizontal=h, vertical='center', wrap_text=wrap)
    if bg: cl.fill = PatternFill('solid', fgColor=bg)
    if bd: cl.border = bdr()
    if nf: cl.number_format = nf
    return cl

def title_row(ws, r, text, ncols, bg=C_DARK):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font  = Font(bold=True, size=13, color=C_WHITE, name='Yu Gothic')
    c.fill  = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 26

def hdr_row(ws, r, headers, bg=C_MID):
    for i, h in enumerate(headers, 1):
        cell(ws, r, i, h, bold=True, bg=bg, fg=C_WHITE)
    ws.row_dimensions[r].height = 20

MAX_HR = 187
DAYS_JP = ['月','火','水','木','金','土','日']

# ──────────────────────────────────────────────────────
#  パース関数
# ──────────────────────────────────────────────────────
def pf(v):
    if v is None: return None
    if isinstance(v, (int,float)): return float(v)
    try: return float(str(v).replace(',',''))
    except: return None

def pdist(v):
    n = pf(v)
    if n is None or n == 0: return 0.0
    return round(n/1000, 3) if n > 50 else round(n, 3)

def ptime(s):
    if s is None: return 0.0
    s = str(s)
    try:
        p = s.split(':')
        if len(p)==3: return round(int(p[0])*60+int(p[1])+float(p[2])/60, 2)
        if len(p)==2: return round(int(p[0])+float(p[1])/60, 2)
    except: return 0.0
    return 0.0

def ppace(s):
    if s in (None,'--',''): return None
    s = str(s)
    p = s.split(':')
    try: return round(int(p[0])+int(p[1])/60, 4)
    except: return None

def fmt_pace(v):
    if v is None: return '—'
    return f"{int(v)}:{int((v%1)*60):02d}"

def zone_of(hr):
    if hr is None: return '—'
    if hr < MAX_HR*0.60: return 'Zone1'
    if hr < MAX_HR*0.70: return 'Zone2'
    if hr < MAX_HR*0.80: return 'Zone3'
    if hr < MAX_HR*0.90: return 'Zone4'
    return 'Zone5'

def time_block(dt):
    h = dt.hour
    if 4 <= h < 12:  return 'AM'
    if 12 <= h < 18: return 'PM'
    return 'Night'

def week_start(dt):
    return (dt - timedelta(days=dt.weekday())).strftime('%Y-%m-%d')

# ──────────────────────────────────────────────────────
#  セッションリスト構築 (rawデータから)
# ──────────────────────────────────────────────────────
C_TYPE=0; C_DATE=1; C_TITLE=3; C_DIST=4; C_CAL=5; C_TIME=6
C_AVG_HR=7; C_MAX_HR=8; C_AVG_PACE=12; C_ELEV=14

def build_sessions(rows):
    sess = []
    for r in rows:
        try: dt = datetime.strptime(r[C_DATE], '%Y-%m-%d %H:%M:%S')
        except: continue
        hr   = pf(r[C_AVG_HR]) if r[C_AVG_HR] != '--' else None
        mhr  = pf(r[C_MAX_HR]) if r[C_MAX_HR] != '--' else None
        dist = pdist(r[C_DIST])
        cal  = pf(r[C_CAL]) or 0
        tmin = ptime(r[C_TIME])
        pace = ppace(r[C_AVG_PACE])
        elev = pf(r[C_ELEV]) if r[C_ELEV] != '--' else None
        sess.append({
            'type':r[C_TYPE], 'dt':dt, 'title':r[C_TITLE],
            'dist':dist, 'cal':cal, 'time':tmin,
            'avg_hr':hr, 'max_hr':mhr,
            'pace':pace, 'zone':zone_of(hr),
            'week':week_start(dt), 'month':dt.strftime('%Y-%m'),
            'elev':elev,
        })
    return sorted(sess, key=lambda s: s['dt'], reverse=True)

sessions_all  = build_sessions(garmin_rows)
sessions      = [s for s in sessions_all if s['type'] != '瞑想']


# ══════════════════════════════════════════════════════
#  Sheet 1: Garminデータ（貼付用）
# ══════════════════════════════════════════════════════
ws_g = wb.active
ws_g.title = 'Garminデータ'

# 案内行
ws_g.merge_cells('A1:H1')
c = ws_g.cell(row=1, column=1,
    value='【使い方】GarminからダウンロードしたCSVを開き、ヘッダー行(1行目)を含めて全行をコピーし、2行目A列に貼り付けてください。その後 更新.bat を実行してください。')
c.font  = Font(bold=True, size=10, color='7B5C00', name='Yu Gothic')
c.fill  = PatternFill('solid', fgColor='FFF3CD')
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws_g.row_dimensions[1].height = 36

# ヘッダー行
for ci, h in enumerate(garmin_header, 1):
    cell(ws_g, 2, ci, h, bold=True, bg=C_MID, fg=C_WHITE, sz=9)

# データ行（既存データ）
for ri, row in enumerate(garmin_rows, 3):
    alt = C_ALT if ri % 2 == 1 else C_WHITE
    for ci, v in enumerate(row, 1):
        cell(ws_g, ri, ci, v, bg=alt, sz=9)

# 列幅
col_widths = [16,22,8,28,8,8,12,10,10,10,10,10,10,10,8,8,8,8,8,8,16,14,12,8,10,10,10,8,8,8,8,8,12,6,8,8,8,8,8,8,8,8,8,12,12,8,8]
for i, w in enumerate(col_widths, 1):
    ws_g.column_dimensions[get_column_letter(i)].width = w

ws_g.freeze_panes = 'A3'

print(f"Garminデータシート: {len(garmin_rows)} 行")


# ══════════════════════════════════════════════════════
#  Sheet 2: ログ（分析用ビュー）
# ══════════════════════════════════════════════════════
ws_log = wb.create_sheet('ログ')

LOG_H = ['No','日付','曜日','時間帯','種別','距離(km)','タイム(分)','平均心拍','最大心拍',
         'ペース(分/km)','ゾーン','消費cal','上昇量(m)','RPE','メモ']
LOG_W = [5,13,5,8,16,10,10,10,10,13,8,10,10,6,30]

title_row(ws_log, 1, 'トレーニングログ（自動生成）', len(LOG_H))
hdr_row(ws_log, 2, LOG_H)
ws_log.freeze_panes = 'A3'
for i,w in enumerate(LOG_W,1):
    ws_log.column_dimensions[get_column_letter(i)].width = w

for ri, s in enumerate(sessions_all, 1):
    r = ri + 2
    alt = C_ALT if ri % 2 == 0 else C_WHITE
    zbg = ZONE_BG.get(s['zone'], C_WHITE)
    zfg = ZONE_FG.get(s['zone'], C_DARK)
    data = [
        ri,
        s['dt'].strftime('%Y/%m/%d'),
        DAYS_JP[s['dt'].weekday()],
        time_block(s['dt']),
        s['type'],
        round(s['dist'],2) if s['dist'] else None,
        round(s['time'],1) if s['time'] else None,
        int(s['avg_hr']) if s['avg_hr'] else None,
        int(s['max_hr']) if s['max_hr'] else None,
        fmt_pace(s['pace']),
        s['zone'],
        int(s['cal']) if s['cal'] else None,
        int(s['elev']) if s['elev'] else None,
        '', '',
    ]
    for ci, v in enumerate(data, 1):
        bg = zbg if ci == 11 else alt
        fg = zfg if ci == 11 else C_DARK
        ah = 'left' if ci in (5,10,11,15) else 'center'
        cell(ws_log, r, ci, v, bg=bg, fg=fg, h=ah)


# ══════════════════════════════════════════════════════
#  集計ヘルパー
# ══════════════════════════════════════════════════════
RUN_TYPES = {'ラン','トラックラン','トレッドミル','屋内ラン'}
ZCOLS_HEX = ['2D6A4F','1D3557','7B5C00','9B3B00','7D0000']
ZONE_NAMES = ['Zone1','Zone2','Zone3','Zone4','Zone5']

def agg(rows):
    run = [s for s in rows if s['type'] in RUN_TYPES]
    zc  = defaultdict(int)
    tc  = defaultdict(int)
    for s in rows:
        zc[s['zone']] += 1
        tc[s['type']] += 1
    return {
        'n':    len(rows),
        'dist': round(sum(s['dist'] for s in run), 1),
        'time': round(sum(s['time'] for s in rows), 0),
        'cal':  int(sum(s['cal']  for s in rows)),
        'zone': dict(zc),
        'type': dict(tc),
    }

# 週別
all_weeks = sorted(set(s['week'] for s in sessions))
weekly = {w: agg([s for s in sessions if s['week']==w]) for w in all_weeks}

# 月別
all_months = sorted(set(s['month'] for s in sessions))
monthly = {m: agg([s for s in sessions if s['month']==m]) for m in all_months}


# ══════════════════════════════════════════════════════
#  Sheet 3: 週間集計
# ══════════════════════════════════════════════════════
ws_wk = wb.create_sheet('週間集計')

WK_H = ['週（月曜）','週末（日曜）','セッション数','走行距離(km)','運動時間(分)','消費cal',
        'Zone1','Zone2','Zone3','Zone4','Zone5']
WK_W = [14,14,12,14,13,12,8,8,8,8,8]

title_row(ws_wk, 1, '週間トレーニング集計', len(WK_H))
hdr_row(ws_wk, 2, WK_H)
ws_wk.freeze_panes = 'A3'
for i,w in enumerate(WK_W,1):
    ws_wk.column_dimensions[get_column_letter(i)].width = w

for ri, wk in enumerate(sorted(all_weeks, reverse=True), 1):
    r  = ri + 2
    alt = C_ALT if ri % 2 == 0 else C_WHITE
    wk_dt = datetime.strptime(wk, '%Y-%m-%d')
    d = weekly[wk]; z = d['zone']
    row_data = [
        wk_dt.strftime('%Y/%m/%d'),
        (wk_dt+timedelta(days=6)).strftime('%Y/%m/%d'),
        d['n'], d['dist'], d['time'], d['cal'],
        z.get('Zone1',0), z.get('Zone2',0),
        z.get('Zone3',0), z.get('Zone4',0), z.get('Zone5',0),
    ]
    for ci, v in enumerate(row_data, 1):
        if 7 <= ci <= 11 and v:
            bg = ZONE_BG.get(f'Zone{ci-6}', alt)
            fg = ZONE_FG.get(f'Zone{ci-6}', C_DARK)
        else:
            bg = alt; fg = C_DARK
        cell(ws_wk, r, ci, v, bg=bg, fg=fg)

de = len(all_weeks) + 2  # data end row

# 合計行
ws_wk.cell(row=de+1, column=1, value='合 計').font = sty(bold=True)
for ci,fn in [(3,sum),(4,sum),(5,sum),(6,sum)]:
    vals = [weekly[w][{3:'n',4:'dist',5:'time',6:'cal'}[ci]] for w in all_weeks]
    cell(ws_wk, de+1, ci, round(sum(vals),1), bold=True, bg='CCCCCC')

# ── グラフ ──
c_ref = Reference(ws_wk, min_col=1, min_row=3, max_row=de)

def bar(title, col, color, anchor, width=22, height=11):
    ch = BarChart(); ch.type='col'; ch.style=10
    ch.title=title; ch.y_axis.title=''; ch.width=width; ch.height=height
    d = Reference(ws_wk, min_col=col, min_row=2, max_row=de)
    ch.add_data(d, titles_from_data=True)
    ch.set_categories(c_ref)
    ch.series[0].graphicalProperties.solidFill = color
    ws_wk.add_chart(ch, anchor)

bar('週別 走行距離 (km)',    4, '4A6FA5', 'M2')
bar('週別 セッション数',      3, '59A14F', 'M17')
bar('週別 消費カロリー',      6, 'E15759', 'M32')

# ゾーン積み上げ
ch_z = BarChart(); ch_z.type='col'; ch_z.grouping='stacked'; ch_z.overlap=100
ch_z.title='週別 ゾーン分布'; ch_z.style=10; ch_z.width=22; ch_z.height=11
for zi, ci in enumerate(range(7,12)):
    ref = Reference(ws_wk, min_col=ci, min_row=2, max_row=de)
    ch_z.add_data(ref, titles_from_data=True)
    ch_z.series[zi].graphicalProperties.solidFill = ZCOLS_HEX[zi]
ch_z.set_categories(c_ref)
ws_wk.add_chart(ch_z, 'M47')

# 週別 円グラフ（直近1週間）
latest_wk = sorted(all_weeks)[-1]
lw_data = weekly[latest_wk]['zone']

ws_wk.cell(row=2, column=26, value='直近週ゾーン').font = sty(bold=True, fg=C_WHITE)
ws_wk.cell(row=2, column=26).fill = fill(C_MID)
pie_start = 3
for pzi, zn in enumerate(ZONE_NAMES):
    cell(ws_wk, pie_start+pzi, 26, zn,  bg=ZONE_BG.get(zn,'E8E8E8'), fg=ZONE_FG.get(zn,C_DARK))
    cell(ws_wk, pie_start+pzi, 27, lw_data.get(zn,0), bg=ZONE_BG.get(zn,'E8E8E8'))

pie_lw = PieChart(); pie_lw.title=f'直近週({latest_wk}) ゾーン'; pie_lw.style=10
pie_lw.width=14; pie_lw.height=11
pd_ref  = Reference(ws_wk, min_col=27, min_row=pie_start, max_row=pie_start+4)
pc_ref  = Reference(ws_wk, min_col=26, min_row=pie_start, max_row=pie_start+4)
pie_lw.add_data(pd_ref); pie_lw.set_categories(pc_ref)
pie_lw.dataLabels = chartlabel.DataLabelList()
pie_lw.dataLabels.showPercent = True; pie_lw.dataLabels.showCatName = True
for idx, hx in enumerate(ZCOLS_HEX):
    pt = DataPoint(idx=idx); pt.graphicalProperties.solidFill = hx
    pie_lw.series[0].dPt.append(pt)
ws_wk.add_chart(pie_lw, 'Z2')


# ══════════════════════════════════════════════════════
#  Sheet 4: 月別集計
# ══════════════════════════════════════════════════════
ws_mo = wb.create_sheet('月別集計')

MO_H = ['月','セッション数','走行距離(km)','運動時間(h)','消費cal',
        'Zone1','Zone2','Zone3','Zone4','Zone5',
        'ラン','TM','バイク系','筋トレ','トラック']
MO_W = [10,12,14,12,12,8,8,8,8,8,8,8,10,8,10]

title_row(ws_mo, 1, '月別トレーニング集計', len(MO_H))
hdr_row(ws_mo, 2, MO_H)
ws_mo.freeze_panes = 'A3'
for i,w in enumerate(MO_W,1):
    ws_mo.column_dimensions[get_column_letter(i)].width = w

for ri, mo in enumerate(sorted(all_months, reverse=True), 1):
    r = ri + 2; alt = C_ALT if ri%2==0 else C_WHITE
    d = monthly[mo]; z = d['zone']; tp = d['type']
    row_data = [
        mo, d['n'], d['dist'], round(d['time']/60,1), d['cal'],
        z.get('Zone1',0), z.get('Zone2',0), z.get('Zone3',0), z.get('Zone4',0), z.get('Zone5',0),
        tp.get('ラン',0)+tp.get('屋内ラン',0),
        tp.get('トレッドミル',0),
        tp.get('屋内バイク',0)+tp.get('Xトレーナー',0)+tp.get('バイク',0),
        tp.get('筋力トレーニング',0),
        tp.get('トラックラン',0),
    ]
    for ci, v in enumerate(row_data, 1):
        if 6 <= ci <= 10 and v:
            bg = ZONE_BG.get(f'Zone{ci-5}', alt); fg = ZONE_FG.get(f'Zone{ci-5}', C_DARK)
        else:
            bg = alt; fg = C_DARK
        cell(ws_mo, r, ci, v, bg=bg, fg=fg)

me = len(all_months) + 2
mc_ref = Reference(ws_mo, min_col=1, min_row=3, max_row=me)

# 月別走行距離折れ線
ch_ml = LineChart(); ch_ml.title='月別 走行距離推移 (km)'; ch_ml.style=10
ch_ml.width=22; ch_ml.height=12
ml_d = Reference(ws_mo, min_col=3, min_row=2, max_row=me)
ch_ml.add_data(ml_d, titles_from_data=True); ch_ml.set_categories(mc_ref)
ch_ml.series[0].graphicalProperties.line.solidFill = '4A6FA5'
ch_ml.series[0].graphicalProperties.line.width = 22000
ws_mo.add_chart(ch_ml, 'Q2')

# 月別ゾーン積み上げ
ch_mz = BarChart(); ch_mz.type='col'; ch_mz.grouping='stacked'; ch_mz.overlap=100
ch_mz.title='月別 ゾーン分布'; ch_mz.style=10; ch_mz.width=22; ch_mz.height=12
for zi, ci in enumerate(range(6,11)):
    ref = Reference(ws_mo, min_col=ci, min_row=2, max_row=me)
    ch_mz.add_data(ref, titles_from_data=True)
    ch_mz.series[zi].graphicalProperties.solidFill = ZCOLS_HEX[zi]
ch_mz.set_categories(mc_ref)
ws_mo.add_chart(ch_mz, 'Q18')

# 月別 種別円グラフ
tot = agg(sessions)['type']
type_items = [(k,v) for k,v in sorted(tot.items(), key=lambda x:-x[1])]
pie_start2 = 3
cell(ws_mo, 2, 20, '種別', bold=True, bg=C_MID, fg=C_WHITE)
cell(ws_mo, 2, 21, '回数', bold=True, bg=C_MID, fg=C_WHITE)
for i,(k,v) in enumerate(type_items):
    cell(ws_mo, pie_start2+i, 20, k)
    cell(ws_mo, pie_start2+i, 21, v)

pie_t = PieChart(); pie_t.title='全体 種別割合'; pie_t.style=10
pie_t.width=16; pie_t.height=13
tp_d = Reference(ws_mo, min_col=21, min_row=pie_start2, max_row=pie_start2+len(type_items)-1)
tp_c = Reference(ws_mo, min_col=20, min_row=pie_start2, max_row=pie_start2+len(type_items)-1)
pie_t.add_data(tp_d); pie_t.set_categories(tp_c)
pie_t.dataLabels = chartlabel.DataLabelList()
pie_t.dataLabels.showPercent = True; pie_t.dataLabels.showCatName = True
PIE_COLORS = ['4A6FA5','F28E2B','E15759','76B7B2','59A14F','EDC948','B07AA1','FF9DA7','9C755F','BAB0AC']
for idx in range(len(type_items)):
    pt = DataPoint(idx=idx); pt.graphicalProperties.solidFill = PIE_COLORS[idx%len(PIE_COLORS)]
    pie_t.series[0].dPt.append(pt)
ws_mo.add_chart(pie_t, 'Q34')


# ══════════════════════════════════════════════════════
#  Sheet 5: ゾーン集計＋全体円グラフ
# ══════════════════════════════════════════════════════
ws_zn = wb.create_sheet('ゾーン集計')

ZN_H = ['ゾーン','心拍範囲','セッション数','割合(%)','合計時間(分)','消費cal']
ZN_W = [20,16,14,12,14,12]

title_row(ws_zn, 1, f'心拍ゾーン別集計（最大心拍 {MAX_HR} bpm）', len(ZN_H))
hdr_row(ws_zn, 2, ZN_H)
for i,w in enumerate(ZN_W,1):
    ws_zn.column_dimensions[get_column_letter(i)].width = w

zone_all = defaultdict(lambda: {'n':0,'time':0,'cal':0})
for s in sessions:
    z = s['zone']
    zone_all[z]['n']    += 1
    zone_all[z]['time'] += s['time']
    zone_all[z]['cal']  += s['cal']

total_n = len(sessions)
ZONES_DEF = [
    ('Zone1', int(MAX_HR*0.50), int(MAX_HR*0.60)-1),
    ('Zone2', int(MAX_HR*0.60), int(MAX_HR*0.70)-1),
    ('Zone3', int(MAX_HR*0.70), int(MAX_HR*0.80)-1),
    ('Zone4', int(MAX_HR*0.80), int(MAX_HR*0.90)-1),
    ('Zone5', int(MAX_HR*0.90), MAX_HR),
    ('—',     None,             None),
]
for ri, (zn, lo, hi) in enumerate(ZONES_DEF, 1):
    r = ri + 2
    rng = f'{lo}–{hi} bpm' if lo else '計測なし'
    bg  = ZONE_BG.get(zn,'E8E8E8'); fg = ZONE_FG.get(zn, C_DARK)
    d   = zone_all[zn]
    pct = round(d['n']/total_n*100, 1) if total_n else 0
    for ci, v in enumerate([zn, rng, d['n'], pct, round(d['time'],0), int(d['cal'])], 1):
        cell(ws_zn, r, ci, v, bg=bg, fg=fg)

# 合計行
tr = len(ZONES_DEF) + 3
cell(ws_zn, tr, 1, '合 計', bold=True, bg='CCCCCC')
for ci, v in [(2,'—'),(3,total_n),(4,100.0),(5,round(sum(d['time'] for d in zone_all.values()),0)),(6,int(sum(d['cal'] for d in zone_all.values())))]:
    cell(ws_zn, tr, ci, v, bold=True, bg='CCCCCC')

# 全体ゾーン円グラフ
pie_z = PieChart(); pie_z.title='全体 ゾーン別割合'; pie_z.style=10
pie_z.width=16; pie_z.height=14
pzd = Reference(ws_zn, min_col=3, min_row=2, max_row=2+len(ZONES_DEF)-1)
pzc = Reference(ws_zn, min_col=1, min_row=3, max_row=2+len(ZONES_DEF)-1)
pie_z.add_data(pzd, titles_from_data=True); pie_z.set_categories(pzc)
pie_z.dataLabels = chartlabel.DataLabelList()
pie_z.dataLabels.showPercent = True; pie_z.dataLabels.showCatName = True
for idx, hx in enumerate(ZCOLS_HEX+['888888']):
    pt = DataPoint(idx=idx); pt.graphicalProperties.solidFill = hx
    pie_z.series[0].dPt.append(pt)
ws_zn.add_chart(pie_z, 'H2')


# ══════════════════════════════════════════════════════
#  Sheet 6: ゾーン定義（参照用）
# ══════════════════════════════════════════════════════
ws_def = wb.create_sheet('ゾーン定義')
DEF_H = ['ゾーン','下限(bpm)','上限(bpm)','最大心拍比','目的']
DEF_W = [12,12,12,14,36]
title_row(ws_def, 1, f'心拍数ゾーン定義（最大心拍数 {MAX_HR} bpm）', len(DEF_H))
hdr_row(ws_def, 2, DEF_H)
for i,w in enumerate(DEF_W,1):
    ws_def.column_dimensions[get_column_letter(i)].width = w
PURPOSES = ['積極的回復・疲労抜き','脂肪燃焼・有酸素基礎','有酸素能力向上','乳酸閾値トレーニング','最大酸素摂取量・スピード']
PCT_LBL  = ['50–60 %','60–70 %','70–80 %','80–90 %','90–100 %']
for ri, (zn, lo, hi) in enumerate(ZONES_DEF[:5], 1):
    r = ri + 2
    bg = ZONE_BG[zn]; fg = ZONE_FG[zn]
    for ci, v in enumerate([zn, lo, hi, PCT_LBL[ri-1], PURPOSES[ri-1]], 1):
        cell(ws_def, r, ci, v, bg=bg, fg=fg, h='left' if ci==5 else 'center')


# ══════════════════════════════════════════════════════
#  シート順序
# ══════════════════════════════════════════════════════
order = ['Garminデータ','ログ','週間集計','月別集計','ゾーン集計','ゾーン定義']
for i, name in enumerate(order):
    wb.move_sheet(name, offset=wb.index(wb[name])*-1 + i)

# ══════════════════════════════════════════════════════
#  保存
# ══════════════════════════════════════════════════════
out = 'hr_zone_analysis.xlsx'
wb.save(out)
print(f'\n✓ 保存完了: {out}')
print(f'  Garminデータ : {len(garmin_rows)} 行')
print(f'  ログ         : {len(sessions_all)} 件')
print(f'  週間集計     : {len(all_weeks)} 週')
print(f'  月別集計     : {len(all_months)} ヶ月')
