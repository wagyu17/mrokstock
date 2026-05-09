import csv
from collections import defaultdict
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── ゾーン定義（最大心拍数 187 bpm 基準） ──────────────────────────
MAX_HR = 187
ZONES = [
    ("Zone 1 – 回復",       int(MAX_HR * 0.50), int(MAX_HR * 0.60),  "AEF0D4", "2D6A4F"),
    ("Zone 2 – 有酸素",     int(MAX_HR * 0.60), int(MAX_HR * 0.70),  "BDE0FB", "1D3557"),
    ("Zone 3 – 有酸素強度", int(MAX_HR * 0.70), int(MAX_HR * 0.80),  "FFF3B0", "7B5C00"),
    ("Zone 4 – 閾値",       int(MAX_HR * 0.80), int(MAX_HR * 0.90),  "FFCBA4", "9B3B00"),
    ("Zone 5 – 無酸素",     int(MAX_HR * 0.90), MAX_HR,              "FFADAD", "7D0000"),
    ("Zone 外（低）",        0,                  int(MAX_HR * 0.50)-1,"E8E8E8", "555555"),
    ("Zone 外（計測なし）",  -1,                 -1,                  "D0D0D0", "333333"),
]

def classify(avg_hr_str):
    if avg_hr_str in ("--", "", None):
        return "Zone 外（計測なし）"
    try:
        hr = int(avg_hr_str.replace(",", ""))
    except ValueError:
        return "Zone 外（計測なし）"
    if hr < int(MAX_HR * 0.50):
        return "Zone 外（低）"
    for name, lo, hi, *_ in ZONES[:5]:
        if lo <= hr <= hi:
            return name
    return "Zone 5 – 無酸素"   # 上振れ保険

# ── CSV 読み込み ────────────────────────────────────────────────────
with open("training_log_2026.csv", "r", encoding="utf-8-sig") as f:
    reader = csv.reader(f)
    header = next(reader)
    rows = list(reader)

# 瞑想除外
rows = [r for r in rows if r[0] != "瞑想"]

# 列インデックス
COL_TYPE  = header.index("アクティビティタイプ")
COL_DATE  = header.index("日付")
COL_TITLE = header.index("タイトル")
COL_AVG_HR= header.index("平均心拍数")
COL_MAX_HR= header.index("最大心拍数")
COL_TIME  = header.index("タイム")
COL_DIST  = header.index("距離")
COL_CAL   = header.index("カロリー")

# ゾーン集計
zone_count   = defaultdict(int)
zone_time    = defaultdict(float)   # 分
zone_cal     = defaultdict(float)
zone_details = defaultdict(list)

for row in rows:
    z = classify(row[COL_AVG_HR])
    zone_count[z] += 1

    # タイム → 分
    t = row[COL_TIME]
    mins = 0.0
    try:
        parts = t.split(":")
        if len(parts) == 3:
            mins = int(parts[0]) * 60 + int(parts[1]) + float(parts[2]) / 60
        elif len(parts) == 2:
            mins = int(parts[0]) + float(parts[1]) / 60
    except:
        pass
    zone_time[z] += mins

    cal = row[COL_CAL]
    try:
        zone_cal[z] += float(cal.replace(",", "")) if cal not in ("--", "") else 0
    except:
        pass

    zone_details[z].append(row)

# ── Excel 作成 ───────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════
# シート1: ゾーン定義
# ═══════════════════════════════════════════
ws_def = wb.active
ws_def.title = "ゾーン定義"

def border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def hfill(hex_bg, hex_fg="000000"):
    return PatternFill("solid", fgColor=hex_bg)

def hfont(bold=False, size=11, color="000000"):
    return Font(bold=bold, size=size, color=color, name="Yu Gothic")

ws_def.column_dimensions["A"].width = 26
ws_def.column_dimensions["B"].width = 14
ws_def.column_dimensions["C"].width = 14
ws_def.column_dimensions["D"].width = 14

# タイトル行
ws_def.merge_cells("A1:D1")
ws_def["A1"] = f"心拍数ゾーン定義（最大心拍数 {MAX_HR} bpm）"
ws_def["A1"].font = hfont(bold=True, size=13, color="FFFFFF")
ws_def["A1"].fill = PatternFill("solid", fgColor="2B3A55")
ws_def["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_def.row_dimensions[1].height = 28

headers_def = ["ゾーン名", "下限 (bpm)", "上限 (bpm)", "目安（最大心拍比）"]
for c, h in enumerate(headers_def, 1):
    cell = ws_def.cell(row=2, column=c, value=h)
    cell.font = hfont(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4A6FA5")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border()

pct_labels = ["50–60 %", "60–70 %", "70–80 %", "80–90 %", "90–100 %", "< 50 %", "—"]
for i, (name, lo, hi, bg, fg) in enumerate(ZONES):
    r = i + 3
    data = [name, lo if lo >= 0 else "—", hi if hi >= 0 else "—", pct_labels[i]]
    for c, v in enumerate(data, 1):
        cell = ws_def.cell(row=r, column=c, value=v)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=(c == 1), color=fg, name="Yu Gothic")
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
        cell.border = border()

# ═══════════════════════════════════════════
# シート2: 集計＋円グラフ
# ═══════════════════════════════════════════
ws_sum = wb.create_sheet("集計・円グラフ")

ws_sum.column_dimensions["A"].width = 26
ws_sum.column_dimensions["B"].width = 12
ws_sum.column_dimensions["C"].width = 16
ws_sum.column_dimensions["D"].width = 16
ws_sum.column_dimensions["E"].width = 16

ws_sum.merge_cells("A1:E1")
ws_sum["A1"] = "心拍数ゾーン別トレーニング集計"
ws_sum["A1"].font = hfont(bold=True, size=13, color="FFFFFF")
ws_sum["A1"].fill = PatternFill("solid", fgColor="2B3A55")
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 28

sum_headers = ["ゾーン名", "回数", "合計時間 (分)", "合計カロリー", "割合 (%)"]
for c, h in enumerate(sum_headers, 1):
    cell = ws_sum.cell(row=2, column=c, value=h)
    cell.font = hfont(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4A6FA5")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border()

total_sessions = sum(zone_count.values())
zone_order = [z[0] for z in ZONES]

for i, zname in enumerate(zone_order):
    r = i + 3
    cnt  = zone_count.get(zname, 0)
    mins = round(zone_time.get(zname, 0), 1)
    cal  = round(zone_cal.get(zname, 0), 0)
    pct  = round(cnt / total_sessions * 100, 1) if total_sessions else 0
    bg   = ZONES[i][3]
    fg   = ZONES[i][4]

    for c, v in enumerate([zname, cnt, mins, int(cal), pct], 1):
        cell = ws_sum.cell(row=r, column=c, value=v)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=(c == 1), color=fg, name="Yu Gothic")
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left")
        cell.border = border()

# 合計行
total_row = len(zone_order) + 3
ws_sum.cell(row=total_row, column=1, value="合計").font = hfont(bold=True)
ws_sum.cell(row=total_row, column=2, value=total_sessions).font = hfont(bold=True)
ws_sum.cell(row=total_row, column=3, value=round(sum(zone_time.values()),1)).font = hfont(bold=True)
ws_sum.cell(row=total_row, column=4, value=int(sum(zone_cal.values()))).font = hfont(bold=True)
ws_sum.cell(row=total_row, column=5, value=100.0).font = hfont(bold=True)
for c in range(1, 6):
    ws_sum.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="CCCCCC")
    ws_sum.cell(row=total_row, column=c).border = border()
    ws_sum.cell(row=total_row, column=c).alignment = Alignment(horizontal="center" if c > 1 else "left")

# 円グラフ
chart = PieChart()
chart.title = "ゾーン別セッション割合"
chart.style = 10

labels_ref = Reference(ws_sum, min_col=1, min_row=3, max_row=3 + len(zone_order) - 1)
data_ref   = Reference(ws_sum, min_col=2, min_row=2, max_row=2 + len(zone_order))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(labels_ref)
chart.dataLabels = openpyxl.chart.label.DataLabelList()
chart.dataLabels.showPercent = True
chart.dataLabels.showCatName = True
chart.dataLabels.showVal = False
chart.dataLabels.separator = "\n"

CHART_COLORS = ["2D6A4F", "1D3557", "7B5C00", "9B3B00", "7D0000", "888888", "555555"]
for idx, color_hex in enumerate(CHART_COLORS):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color_hex
    chart.series[0].dPt.append(pt)

chart.width  = 18
chart.height = 14
ws_sum.add_chart(chart, "G2")

# ═══════════════════════════════════════════
# シート3〜: ゾーン別詳細
# ═══════════════════════════════════════════
detail_cols = ["アクティビティタイプ", "日付", "タイトル", "平均心拍数", "最大心拍数", "タイム", "距離", "カロリー"]
detail_idx  = [header.index(c) for c in detail_cols]

for zname, bg, fg in [(z[0], z[3], z[4]) for z in ZONES]:
    detail_rows = zone_details.get(zname, [])
    if not detail_rows:
        continue

    sheet_name = zname[:31]
    ws_d = wb.create_sheet(sheet_name)

    ws_d.merge_cells(f"A1:{get_column_letter(len(detail_cols))}1")
    ws_d["A1"] = f"{zname}  （{len(detail_rows)} セッション）"
    ws_d["A1"].font = Font(bold=True, size=12, color="FFFFFF", name="Yu Gothic")
    ws_d["A1"].fill = PatternFill("solid", fgColor="2B3A55")
    ws_d["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_d.row_dimensions[1].height = 24

    for c, h in enumerate(detail_cols, 1):
        cell = ws_d.cell(row=2, column=c, value=h)
        cell.font = hfont(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4A6FA5")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border()

    for r_i, row in enumerate(detail_rows, 3):
        for c_i, idx in enumerate(detail_idx, 1):
            cell = ws_d.cell(row=r_i, column=c_i, value=row[idx])
            cell.fill = PatternFill("solid", fgColor=bg if r_i % 2 == 1 else "FFFFFF")
            cell.font = Font(color=fg if r_i % 2 == 1 else "333333", name="Yu Gothic")
            cell.alignment = Alignment(horizontal="center" if c_i > 1 else "left")
            cell.border = border()

    col_widths = [22, 22, 30, 14, 14, 14, 12, 12]
    for c, w in enumerate(col_widths, 1):
        ws_d.column_dimensions[get_column_letter(c)].width = w

# ─── 保存 ────────────────────────────────────────────────────────────
out = "hr_zone_analysis.xlsx"
wb.save(out)
print(f"保存完了: {out}")
